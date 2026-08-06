# -*- coding: utf-8 -*-
"""
真实测试结果 -> Excel 导出器

读取 conftest 在 sessionfinish dump 的 reports/output/test_results.json,
生成 8 列 Excel(复用 VLAN 更新版样式)，并把每个测试步骤的完整详情
写入独立的“步骤明细”sheet，人工复验命令写入“复验命令”sheet。
超长文本会无损分片，避免触发 Excel 单元格 32767 字符上限。

内容来自真实测试执行: 每步标题+状态+SSH验证输出、用例真实 PASS/FAIL、
失败错误信息、失败截图文件路径。比手写 YAML 用例更全面、更真实。

用法:
  python utils/test_results_to_excel.py                       # 读最新 test_results.json
  python utils/test_results_to_excel.py -i xxx.json -o y.xlsx  # 指定输入输出
"""
import os
import json
import logging
import argparse
import re

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    _HAS_OPENPYXL = False

from utils.step_recorder import redact_sensitive_text as _registry_redact_sensitive_text

# 8 列表头与列宽(与 VLAN 更新版对齐; 内联以保持本模块独立可运行)
HEADERS = ["模块", "测试项", "前提条件", "测试场景", "测试步骤", "预期结果", "测试结果", "备注"]
COL_WIDTHS = [28, 22, 55, 45, 55, 60, 12, 35]
_COL_ALIGN = {
    1: ("center", "center"), 2: ("center", "center"),
    3: ("left", "top"), 4: ("left", "top"), 5: ("left", "top"),
    6: ("left", "top"), 7: ("left", "top"), 8: ("left", "top"),
}


def _build_styles():
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Calibri", size=11, bold=False)
    data_aligns = {c: Alignment(horizontal=h, vertical=v, wrap_text=True)
                   for c, (h, v) in _COL_ALIGN.items()}
    return border, header_fill, header_font, header_align, data_font, data_aligns

_HERE = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_HERE)
_DEFAULT_JSON = os.path.join(_PROJECT_ROOT, "reports", "output", "test_results.json")

# module_key -> 中文名(从 original_name 推 key 后查)
_MODULE_NAMES = {
    "vlan": "VLAN设置", "nat_rule": "NAT规则", "port_map": "端口映射", "dmz_host": "DMZ主机",
    "upnp_setting": "UPnP/NAT设置", "static_route": "静态路由", "cross_layer_service": "跨三层服务",
    "multi_wan_lb": "多线负载", "protocol_route": "协议分流", "port_route": "端口分流",
    "domain_route": "域名分流", "updown_route": "上下行分离", "ip_rate_limit": "IP限速",
    "mac_rate_limit": "MAC限速", "stream_control": "智能流控", "ipv6_static": "IPv6前缀静态分配",
    "custom_protocol": "自定义协议", "advanced_custom_protocol": "高级自定义协议",
    "dhcp_server": "DHCP服务端", "dhcp_static": "DHCP静态分配", "dhcp_lease": "DHCP客户端",
    "dhcp_acl_mac": "DHCP黑白名单", "dns_accelerate": "DNS加速服务", "dns_multi_line": "多线路DNS服务",
    "igmp_proxy": "IGMP代理", "iptv": "IPTV透传", "udp_proxy": "UDPXY设置",
    "ip_group": "IP分组", "mac_group": "MAC分组", "port_group": "端口分组",
    "domain_group": "域名分组", "time_plan": "时间计划", "protocol_group": "协议分组",
    "advanced": "安全中心-高级设置", "acl": "安全中心-ACL规则",
    "conn_limit": "安全中心-连接数限制", "mac_access_control": "安全中心-MAC访问控制",
    "app_protocol": "安全中心-应用协议控制",
    "url_black": "安全中心-网址浏览控制-网址黑白名单综合测试(L1-L4)",
    "url_black_http_https": "安全中心-网址黑白名单功能测试(HTTP/HTTPS与白名单外链开关L5)",
    "domain_blacklist": "安全中心-网址浏览控制-禁止娱乐网站综合测试(L1-L4)",
    "domain_blacklist_http_https": "安全中心-禁止娱乐网站功能测试(HTTP/HTTPS真实阻断L5)",
    "custom_domain_group": "安全中心-网址浏览控制-自定义网址库综合测试(L1-L2)",
    "custom_domain_group_http_https": "安全中心-自定义网址库功能测试(禁止娱乐网站联动HTTP/HTTPS L5)",
    "ftp_server": "高级服务-本地服务-FTP服务",
    "samba_server": "高级服务-本地服务-Samba服务",
    "http_server": "高级服务-本地服务-HTTP服务",
    "snmp_server": "高级服务-本地服务-SNMP服务",
    "virtual_machine": "高级服务-虚拟机",
    "basic_setting": "设备设置-基础设置",
    "alg_setting": "设备设置-高级管理-ALG设置",
    "protocol_control": "设备设置-高级管理-协议控制",
    "kernel_setting": "设备设置-高级管理-内核设置",
    "account_setting": "设备设置-登录管理-账号设置",
    "cloud_service_binding": "设备设置-云服务绑定",
    "ppp_package": "认证服务-套餐管理",
    "pppuser": "认证服务-账号管理",
    "ppp_passwd": "认证服务-自助密码管理",
    "ppp_paylog": "认证服务-总账管理",
    "coupon": "认证服务-上网码",
}

_STATUS_CN = {
    "passed": "通过",
    "failed": "失败",
    "error": "失败",
    "warning": "警告",
    "not_applicable": "不适用",
    # Historical pytest/report spelling; external reports use the four-state
    # Chinese vocabulary required by the project.
    "skipped": "不适用",
}
_STATUS_MARK = {
    "passed": "✓",
    "failed": "✗",
    "error": "✗",
    "warning": "!",
    "not_applicable": "○",
    "skipped": "○",
}
_EXCEL_CELL_LIMIT = 32767
# 为不同 Excel/WPS 版本和 openpyxl 的内部处理留少量余量。
_EXCEL_CHUNK_SIZE = 32000
_SENSITIVE_TEXT_RE = re.compile(
    r"(?i)(community|auth(?:entication)?[_ -]?(?:pass(?:word)?|key)|"
    r"priv(?:acy)?[_ -]?(?:pass(?:word)?|key)|password|passwd|secret|"
    r"团体名|认证(?:口令|密码|密钥)|隐私(?:口令|密码|密钥))"
    r"\s*[:=]\s*([^,;，\r\n}\]]+)"
)
_SENSITIVE_JSON_RE = re.compile(
    r"(?i)([\"'](?:community|auth(?:entication)?[_ -]?(?:pass(?:word)?|key)|"
    r"priv(?:acy)?[_ -]?(?:pass(?:word)?|key)|password|passwd|secret|"
    r"团体名|认证(?:口令|密码|密钥)|隐私(?:口令|密码|密钥))[\"']\s*:\s*)"
    r"([\"'])(.*?)\2"
)
_SENSITIVE_SNMP_ARG_RE = re.compile(
    r"(?i)\b(?:snmpget|snmpwalk|snmpbulkwalk)\b.*"
    r"(?:^|\s)(?:-c|-A|-X|--community|--auth-pass|--priv-pass)\s+\S+"
)


def _status_cn(status) -> str:
    """Expose only the project's four Chinese report states."""
    return _STATUS_CN.get(str(status or "").strip().lower(), "警告")


def _redact_sensitive_text(value) -> str:
    text = _registry_redact_sensitive_text(value)
    text = _SENSITIVE_JSON_RE.sub(
        lambda match: f'{match.group(1)}"[已隐藏]"', text
    )
    return _SENSITIVE_TEXT_RE.sub(
        lambda match: f"{match.group(1)}=[已隐藏]", text
    )


def _module_key(original_name: str) -> str:
    """test_nat_rule_comprehensive[chromium] -> nat_rule"""
    base = original_name.split("[")[0]
    if base.startswith("test_"):
        base = base[5:]
    for suf in ("_comprehensive_flow", "_comprehensive", "_flow"):
        if base.endswith(suf):
            return base[: -len(suf)]
    return base


def _module_cn(original_name: str, name: str) -> str:
    key = _module_key(original_name)
    return _MODULE_NAMES.get(key, name or key or "未分类")


def _split_excel_text(value, chunk_size: int = _EXCEL_CHUNK_SIZE):
    """无损拆分超长文本，确保每个 Excel 单元格都低于长度上限。"""
    text = "" if value is None else str(value)
    if not text:
        return [""]
    return [text[i:i + chunk_size] for i in range(0, len(text), chunk_size)]


def _fit_excel_cell(value, overflow_hint: str = "完整内容见“步骤明细”sheet") -> str:
    """把单值安全放入一个单元格；完整长文本由步骤明细sheet承载。"""
    text = "" if value is None else str(value)
    if len(text) <= _EXCEL_CELL_LIMIT:
        return text
    suffix = f"\n…（{overflow_hint}）"
    return text[:_EXCEL_CELL_LIMIT - len(suffix)] + suffix


def _neutralize_excel_formulas(workbook):
    """Store every report string as an XLSX string, never as a formula.

    Report content originates in UI labels, API/SSH output, error messages and
    user-entered values.  openpyxl automatically treats a value beginning with
    ``=`` as a formula; spreadsheet applications also recognize other formula
    prefixes in imported content.  Marking every string cell explicitly as a
    string preserves its visible value (including ``=``, ``+``, ``-`` or ``@``)
    while preventing execution.  A workbook-wide final pass also protects new
    report columns added later.
    """
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.data_type = "s"


def _render_steps(steps) -> str:
    """测试步骤列: 紧凑保留每一步的状态、标题和用时。

    SSH输出和逐项断言改由“步骤明细”sheet完整保存。当前25步场景会全部
    出现在本单元格；极端超长用例则只在完整行边界截断并明确指向明细sheet，
    不会出现 openpyxl 在某个步骤中间静默截断的情况。
    """
    lines = []
    for i, st in enumerate(steps or [], 1):
        status = st.get("status", "")
        mark = _STATUS_MARK.get(status, "·")
        title = st.get("name", f"步骤{i}")
        dur = st.get("duration", "")
        head = f"{i}. [{mark}] {title}"
        if dur:
            head += f"  ({dur})"
        lines.append(head)
    rendered = "\n".join(lines)
    if len(rendered) <= _EXCEL_CELL_LIMIT:
        return rendered

    kept = []
    for index, line in enumerate(lines):
        remaining = len(lines) - index
        pointer = f"… 另有 {remaining} 步，完整内容见“步骤明细”sheet"
        candidate = "\n".join(kept + [line, pointer])
        if len(candidate) > _EXCEL_CELL_LIMIT:
            return "\n".join(kept + [pointer])
        kept.append(line)
    return "\n".join(kept)


def _render_result(tc) -> str:
    status = tc.get("status", "")
    text = _status_cn(status)
    err = tc.get("error_message")
    if err:
        text += f"\n错误: {err}"
    return _fit_excel_cell(text, "完整错误见“步骤明细”sheet")


def _chunk_value(chunks, index: int, repeat_single: bool = False) -> str:
    """取分片；单片元数据可按行重复，长文本分片则逐行展开。"""
    if repeat_single and len(chunks) == 1:
        return chunks[0]
    return chunks[index] if index < len(chunks) else ""


def _boolean_label(value, *, unknown: str = "未标注") -> str:
    """把 JSON 布尔值或常见字符串转为中文状态。"""
    if isinstance(value, bool):
        return "是" if value else "否"
    if value is None or value == "":
        return unknown
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "on", "y", "是"}:
        return "是"
    if normalized in {"0", "false", "no", "off", "n", "否"}:
        return "否"
    return unknown


def _read_only_label(command) -> str:
    """优先读显式 read_only，否则由 effect 判定是否只读。"""
    if not isinstance(command, dict):
        return "未标注"
    if "read_only" in command:
        explicit = _boolean_label(command.get("read_only"))
        return "只读（不修改配置）" if explicit == "是" else explicit
    effect = str(command.get("effect") or "").strip().lower()
    if not effect:
        return "未标注"
    normalized = effect.replace("-", "_").replace(" ", "_")
    if normalized in {
        "read_only", "readonly", "query", "diagnostic", "no_side_effects",
        "只读", "无副作用",
    }:
        return "只读（不修改配置）"
    return "否"


def _interactive_metadata(command):
    """返回 ``(是否交互, 交互提示)``，兼容布尔值和文本写法。"""
    if not isinstance(command, dict):
        return "未标注", ""
    raw = command.get("interactive")
    hint = (command.get("interactive_hint") or
            command.get("interaction_hint") or "")
    status = _boolean_label(raw)
    if status == "未标注" and raw not in (None, ""):
        status = "是"
        if not hint:
            hint = str(raw)
    elif status == "未标注" and hint:
        status = "是"
    return status, str(hint)


def _write_step_details(ws, data, styles):
    """逐 detail 写行；任何超长字段均拆成多行且不丢失字符。"""
    border, hfill, hfont, halign, dfon, daligns = styles
    headers = [
        "用例", "模块", "步骤序号", "步骤名称", "步骤描述", "步骤状态",
        "步骤耗时", "详情序号", "详情", "步骤实际", "步骤错误", "用例错误",
    ]
    widths = [30, 30, 10, 38, 48, 12, 12, 12, 90, 55, 55, 55]
    for ci, title in enumerate(headers, 1):
        cell = ws.cell(1, ci, title)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = halign
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    row_index = 2
    for tc in data.get("test_cases", []):
        case_chunks = _split_excel_text(tc.get("name", ""))
        module_chunks = _split_excel_text(_module_cn(
            tc.get("original_name", ""), tc.get("name", "")
        ))
        case_error_chunks = _split_excel_text(
            _redact_sensitive_text(tc.get("error_message") or "")
        )
        steps = tc.get("steps", []) or []
        for step_no, step in enumerate(steps, 1):
            name_chunks = _split_excel_text(step.get("name", f"步骤{step_no}"))
            description_chunks = _split_excel_text(step.get("description", ""))
            status_chunks = _split_excel_text(_status_cn(step.get("status", "")))
            duration_chunks = _split_excel_text(step.get("duration", ""))
            actual_chunks = _split_excel_text(
                _redact_sensitive_text(step.get("actual", ""))
            )
            step_error_chunks = _split_excel_text(
                _redact_sensitive_text(step.get("error_message") or "")
            )

            detail_rows = []
            details = step.get("details", []) or []
            if not details:
                detail_rows.append(("", ""))
            for detail_no, detail in enumerate(details, 1):
                detail_chunks = _split_excel_text(_redact_sensitive_text(detail))
                total_parts = len(detail_chunks)
                for part_no, chunk in enumerate(detail_chunks, 1):
                    number = (str(detail_no) if total_parts == 1 else
                              f"{detail_no}.{part_no}/{total_parts}")
                    detail_rows.append((number, chunk))

            row_count = max(
                len(detail_rows), len(case_chunks), len(module_chunks),
                len(name_chunks), len(description_chunks), len(status_chunks),
                len(duration_chunks), len(actual_chunks), len(step_error_chunks),
                len(case_error_chunks), 1,
            )
            for part_index in range(row_count):
                detail_no, detail_text = (
                    detail_rows[part_index]
                    if part_index < len(detail_rows) else ("", "")
                )
                values = [
                    _chunk_value(case_chunks, part_index, repeat_single=True),
                    _chunk_value(module_chunks, part_index, repeat_single=True),
                    step_no,
                    _chunk_value(name_chunks, part_index, repeat_single=True),
                    _chunk_value(description_chunks, part_index, repeat_single=True),
                    _chunk_value(status_chunks, part_index, repeat_single=True),
                    _chunk_value(duration_chunks, part_index, repeat_single=True),
                    detail_no,
                    detail_text,
                    _chunk_value(actual_chunks, part_index),
                    _chunk_value(step_error_chunks, part_index),
                    _chunk_value(case_error_chunks, part_index),
                ]
                for ci, value in enumerate(values, 1):
                    # 所有字符串在进入单元格前均已分片；这里再作防御性保护。
                    safe_value = (_fit_excel_cell(value)
                                  if isinstance(value, str) else value)
                    cell = ws.cell(row_index, ci, safe_value)
                    cell.font = dfon
                    cell.alignment = Alignment(
                        horizontal="center" if ci in {3, 6, 7, 8} else "left",
                        vertical="top", wrap_text=True,
                    )
                    cell.border = border
                row_index += 1


def _write_verification_commands(ws, data, styles):
    """每条人工复验命令独立成行；超长字段按完整字符串无损分片。"""
    border, hfill, hfont, halign, dfon, daligns = styles
    headers = [
        "用例", "步骤序号", "步骤", "命令序号", "目标", "目标类型", "IP", "Shell",
        "用途", "有效时机", "交互", "交互提示", "命令分片", "命令", "预期",
        "实际", "只读", "可复制", "含敏感信息",
    ]
    widths = [30, 10, 38, 10, 16, 14, 18, 14, 40, 36, 10, 42, 12, 90, 48, 48, 22, 10, 14]
    for ci, title in enumerate(headers, 1):
        cell = ws.cell(1, ci, title)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = halign
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    command_font = Font(name="Consolas", size=10, bold=False)
    row_index = 2
    for tc in data.get("test_cases", []):
        case_name = tc.get("name", "")
        steps = tc.get("steps", []) or []
        for step_no, step in enumerate(steps, 1):
            commands = step.get("verification_commands", []) or []
            if isinstance(commands, (str, bytes, dict)):
                commands = [commands]
            for command_no, raw_command in enumerate(commands, 1):
                command = (dict(raw_command) if isinstance(raw_command, dict)
                           else {"command": raw_command})
                contains_secret = (
                    _boolean_label(command.get("contains_secret")) == "是"
                )
                command_text = str(command.get("command", ""))
                if (
                    _SENSITIVE_TEXT_RE.search(command_text) or
                    _SENSITIVE_JSON_RE.search(command_text) or
                    _SENSITIVE_SNMP_ARG_RE.search(command_text)
                ):
                    contains_secret = True
                if contains_secret:
                    # 兼容绕过 StepRecorder 直接构造的旧/外部 JSON：Excel
                    # 展示层再次不可逆隐藏正文，并禁止标为可复制。
                    command["command"] = "[命令已隐藏：包含敏感信息]"
                    command["actual"] = "[命令已隐藏：包含敏感信息]"
                    command["copy_ready"] = False
                interactive_status, interactive_hint = _interactive_metadata(command)
                field_chunks = {
                    "case": _split_excel_text(case_name),
                    "step": _split_excel_text(step.get("name", f"步骤{step_no}")),
                    "target": _split_excel_text(
                        command.get("target_label") or command.get("target", "")
                    ),
                    "target_type": _split_excel_text(command.get("target", "")),
                    "host": _split_excel_text(command.get("host", "")),
                    "shell": _split_excel_text(command.get("shell", "")),
                    "purpose": _split_excel_text(
                        _redact_sensitive_text(command.get("purpose", ""))
                    ),
                    "valid_when": _split_excel_text(
                        _redact_sensitive_text(command.get("valid_when", ""))
                    ),
                    "interactive": _split_excel_text(interactive_status),
                    "interactive_hint": _split_excel_text(
                        _redact_sensitive_text(interactive_hint)
                    ),
                    "command": _split_excel_text(command.get("command", "")),
                    "expected": _split_excel_text(
                        _redact_sensitive_text(command.get("expected", ""))
                    ),
                    "actual": _split_excel_text(
                        _redact_sensitive_text(command.get("actual", ""))
                    ),
                }
                row_count = max(len(chunks) for chunks in field_chunks.values())
                command_part_count = len(field_chunks["command"])
                for part_index in range(row_count):
                    if part_index < command_part_count:
                        command_part = ("1" if command_part_count == 1 else
                                        f"{part_index + 1}/{command_part_count}")
                    else:
                        command_part = ""
                    values = [
                        _chunk_value(field_chunks["case"], part_index, repeat_single=True),
                        step_no,
                        _chunk_value(field_chunks["step"], part_index, repeat_single=True),
                        command_no,
                        _chunk_value(field_chunks["target"], part_index, repeat_single=True),
                        _chunk_value(field_chunks["target_type"], part_index, repeat_single=True),
                        _chunk_value(field_chunks["host"], part_index, repeat_single=True),
                        _chunk_value(field_chunks["shell"], part_index, repeat_single=True),
                        _chunk_value(field_chunks["purpose"], part_index, repeat_single=True),
                        _chunk_value(field_chunks["valid_when"], part_index, repeat_single=True),
                        _chunk_value(field_chunks["interactive"], part_index, repeat_single=True),
                        _chunk_value(field_chunks["interactive_hint"], part_index,
                                     repeat_single=True),
                        command_part,
                        _chunk_value(field_chunks["command"], part_index),
                        _chunk_value(field_chunks["expected"], part_index, repeat_single=True),
                        _chunk_value(field_chunks["actual"], part_index, repeat_single=True),
                        _read_only_label(command),
                        _boolean_label(command.get("copy_ready")),
                        _boolean_label(command.get("contains_secret")),
                    ]
                    for ci, value in enumerate(values, 1):
                        safe_value = (_fit_excel_cell(value, "完整内容见后续分片")
                                      if isinstance(value, str) else value)
                        cell = ws.cell(row_index, ci, safe_value)
                        cell.font = command_font if ci == 14 else dfon
                        cell.alignment = Alignment(
                            horizontal=("center" if ci in {
                                2, 4, 5, 6, 7, 8, 11, 13, 17, 18, 19,
                            }
                                        else "left"),
                            vertical="top",
                            wrap_text=True,
                        )
                        cell.border = border
                    row_index += 1


def _merge_module_col(ws, col: str, first_row: int, keys):
    n = len(keys)
    i = 0
    while i < n:
        j = i
        while j + 1 < n and keys[j + 1] == keys[i] and keys[i]:
            j += 1
        if j > i:
            ws.merge_cells(f"{col}{first_row + i}:{col}{first_row + j}")
        i = j + 1


def _write_summary(ws, data, styles):
    """汇总 sheet: 统计卡片 + 用例简表"""
    border, hfill, hfont, halign, dfon, daligns = styles
    from openpyxl.styles import Font
    ws.cell(1, 1, "测试结果汇总").font = Font(bold=True, size=13)
    stats = [
        ("总计", data.get("total", 0)),
        ("通过", data.get("passed", 0)),
        ("失败", data.get("failed", 0)),
        (
            "警告",
            sum(
                1 for case in data.get("test_cases", [])
                if case.get("status") == "warning"
            ),
        ),
        (
            "不适用",
            sum(
                1 for case in data.get("test_cases", [])
                if case.get("status") in {"not_applicable", "skipped"}
            ),
        ),
        ("总步骤", data.get("total_steps", 0)),
        ("用时", data.get("duration", "")),
        ("开始", data.get("start_time", "")),
        ("结束", data.get("end_time", "")),
    ]
    for i, (k, v) in enumerate(stats, start=3):
        ws.cell(i, 1, k).font = dfon
        ws.cell(i, 2, _fit_excel_cell(v)).font = dfon

    # 用例简表
    headers = ["用例", "模块", "状态", "用时", "步骤数", "错误信息"]
    start = 3 + len(stats) + 1
    for ci, h in enumerate(headers, 1):
        c = ws.cell(start, ci, h)
        c.fill = hfill; c.font = hfont; c.alignment = halign; c.border = border
    for idx, tc in enumerate(data.get("test_cases", []), start=start + 1):
        vals = [
            tc.get("name", ""),
            _module_cn(tc.get("original_name", ""), tc.get("name", "")),
            _status_cn(tc.get("status", "")),
            tc.get("duration", ""),
            tc.get("step_count", 0),
            tc.get("error_message") or "",
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(idx, ci, _fit_excel_cell(v))
            cell.font = dfon
            cell.alignment = daligns[ci]
            cell.border = border
    ws.column_dimensions["A"].width = 30
    for c in "BCDEF":
        ws.column_dimensions[c].width = 18
    ws.freeze_panes = f"A{start + 1}"


def export_results_to_excel(json_path: str, output_path: str):
    """返回 (success: bool, message: str)"""
    if not _HAS_OPENPYXL:
        return False, "缺少 openpyxl 依赖, 请执行: pip install openpyxl"
    if not os.path.exists(json_path):
        return False, f"测试结果文件不存在: {json_path}\n请先运行测试(会自动生成该JSON)"
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        return False, f"读取结果JSON失败: {e}"

    cases = data.get("test_cases", [])
    if not cases:
        return False, "结果JSON中无测试用例"

    wb = Workbook()
    wb.remove(wb.active)
    styles = _build_styles()
    border, hfill, hfont, halign, dfon, daligns = styles

    # 主表: 8 列每用例一行
    ws = wb.create_sheet("测试结果明细")
    for ci, title in enumerate(HEADERS, 1):
        c = ws.cell(1, ci, title)
        c.fill = hfill; c.font = hfont; c.alignment = halign; c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS[ci - 1]
    ws.freeze_panes = "A2"

    module_keys = []
    row = 2
    for tc in cases:
        orig = tc.get("original_name", "")
        name = tc.get("name", "")
        module_cn = _module_cn(orig, name)
        scenario = f"{name}（用时 {tc.get('duration', '')}，共 {tc.get('step_count', 0)} 步）"
        steps_text = _render_steps(tc.get("steps", []))
        result_text = _render_result(tc)
        shot = tc.get("screenshot_path", "")
        remark = f"失败截图: {shot}" if shot else ""
        values = [module_cn, name, "", scenario, steps_text, "", result_text, remark]
        for ci, val in enumerate(values, 1):
            safe_val = _fit_excel_cell(val)
            cell = ws.cell(row, ci, safe_val)
            cell.font = dfon
            cell.alignment = daligns[ci]
            cell.border = border
        module_keys.append(module_cn)
        row += 1
    _merge_module_col(ws, "A", 2, module_keys)

    # 汇总 sheet(放最前)
    ws_summary = wb.create_sheet("汇总", 0)
    _write_summary(ws_summary, data, styles)

    # 完整步骤明细：逐项断言/SSH输出各占一行，超长详情自动无损分片。
    ws_steps = wb.create_sheet("步骤明细")
    _write_step_details(ws_steps, data, styles)

    # 结构化人工复验命令：正常命令一行，超长命令按分片顺序无损展开。
    ws_commands = wb.create_sheet("复验命令")
    _write_verification_commands(ws_commands, data, styles)

    try:
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        _neutralize_excel_formulas(wb)
        wb.save(output_path)
    except Exception as e:
        return False, f"保存失败: {e}"

    return True, (f"已导出 {len(cases)} 条用例 → {output_path}"
                  "（含汇总sheet + 测试结果明细sheet + 步骤明细sheet"
                  " + 复验命令sheet）")


def _main():
    ap = argparse.ArgumentParser(description="把真实测试结果(test_results.json)导出为 Excel")
    ap.add_argument("-i", "--input", default=_DEFAULT_JSON, help="test_results.json 路径")
    ap.add_argument("-o", "--output", default=None, help="输出 xlsx 路径")
    args = ap.parse_args()
    out = args.output or os.path.join(
        _PROJECT_ROOT, "reports",
        f"测试结果_{os.path.basename(os.path.dirname(args.input))}.xlsx",
    )
    ok, msg = export_results_to_excel(args.input, out)
    print(("[OK] " if ok else "[FAIL] ") + msg)
    if not ok:
        raise SystemExit(1)


if __name__ == "__main__":
    _main()
