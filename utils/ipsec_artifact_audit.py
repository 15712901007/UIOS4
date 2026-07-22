"""Audit IPsec JSON/HTML/Excel artifacts for consistency and secret safety."""
from __future__ import annotations

import json
import re
from html.parser import HTMLParser
from pathlib import Path
from typing import Dict, Iterable, List


REQUIRED_SECTIONS = (
    "【测试操作】", "【页面验证】", "【后端验证】",
    "【运行时验证】", "【协议验证】", "【清理结果】",
)
MAC_PATTERN = re.compile(r"(?i)(?:[0-9a-f]{2}:){5}[0-9a-f]{2}")
SENSITIVE_ASSIGNMENT = re.compile(
    r"(?i)(?:password|passwd|secret|psk|token|cookie|private[_ -]?key)"
    r"\s*[=:]\s*(?!<redacted>|\*{3,}|\{configured)[^\s,;]+"
)
UNSAFE_COMMAND = re.compile(
    r"\b(?:if|then|fi|for|do|done|while|case|esac|base64)\b|"
    r"\$\(|\$\{|\$(?:\?|[A-Za-z_][A-Za-z0-9_]*)|"
    r"(?:^|[ ;])(?:rc|[A-Za-z_][A-Za-z0-9_]*)\s*=|"
    r"(?:^|[ ;])\[(?:router|client|peer)\](?:$|[ ;])"
)
MUTATING_COMMAND = re.compile(
    r"(?i)(?:^|[ ;|])(?:rm|mv|cp|kill|reboot|shutdown|service|systemctl)\s|"
    r"\b(?:delete|insert|update|drop|truncate)\b|"
    r"\bip\s+(?:addr|route|rule|link|xfrm)\s+(?:add|del|delete|flush|set)\b|"
    r"\bip\s+xfrm\s+(?:state|policy)\s+flush\b|"
    r"\bswanctl\s+--(?:initiate|terminate|rekey|load)\b"
)


class _CommandParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.commands: List[str] = []
        self._parts: List[str] | None = None

    def handle_starttag(self, tag, attrs):
        if tag.lower() == "code" and self._parts is None:
            element_id = str(dict(attrs).get("id", ""))
            if element_id.startswith("verification-command-"):
                self._parts = []

    def handle_data(self, data):
        if self._parts is not None:
            self._parts.append(data)

    def handle_endtag(self, tag):
        if tag.lower() == "code" and self._parts is not None:
            self.commands.append("".join(self._parts))
            self._parts = None


def _workbook_values(path: Path) -> List[str]:
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return [
            str(cell.value or "")
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        ]
    finally:
        workbook.close()


def audit_ipsec_artifacts(
    json_path, html_path, excel_path,
    sensitive_values: Iterable[str] = (),
) -> Dict[str, int]:
    json_path, html_path, excel_path = map(
        Path, (json_path, html_path, excel_path)
    )
    data = json.loads(json_path.read_text(encoding="utf-8"))
    html_text = html_path.read_text(encoding="utf-8")
    workbook_values = _workbook_values(excel_path)
    excel_text = "\n".join(workbook_values)
    serialized = json.dumps(data, ensure_ascii=False) + html_text + excel_text
    cases = data.get("test_cases", []) or []
    if len(cases) != 1 or "ipsec" not in str(cases[0].get("name", "")).lower():
        raise AssertionError("产物不是唯一IPsec综合用例")
    steps = cases[0].get("steps", []) or []
    if not steps:
        raise AssertionError("IPsec产物缺少步骤")
    allowed_statuses = {"passed", "failed", "warning", "not_applicable", "skipped"}
    for index, step in enumerate(steps, 1):
        if str(step.get("status", "")) not in allowed_statuses:
            raise AssertionError(f"步骤{index}状态不合法")
        detail = "\n".join(map(str, step.get("details", []) or []))
        missing = [marker for marker in REQUIRED_SECTIONS if marker not in detail]
        if missing:
            raise AssertionError(f"步骤{index}缺少六段证据")
    if MAC_PATTERN.search(serialized):
        raise AssertionError("产物包含硬件地址")
    if SENSITIVE_ASSIGNMENT.search(serialized):
        raise AssertionError("产物包含疑似认证明文")
    for value in sensitive_values:
        value = str(value or "")
        if len(value) >= 4 and value in serialized:
            raise AssertionError("产物包含运行时敏感值")
    for value in workbook_values:
        if value.lstrip().startswith(("=", "+", "-", "@")):
            raise AssertionError("Excel包含潜在公式注入")

    commands = [
        item for step in steps
        for item in (step.get("verification_commands", []) or [])
    ]
    expected: List[str] = []
    for index, item in enumerate(commands, 1):
        command = str(item.get("command", ""))
        if item.get("target") not in {"router", "client"}:
            raise AssertionError(f"复验命令{index}目标不合规")
        if not command or "\n" in command or "\r" in command:
            raise AssertionError(f"复验命令{index}不是单行命令")
        if item.get("contains_secret") or not item.get("copy_ready"):
            raise AssertionError(f"复验命令{index}不可安全复制")
        if UNSAFE_COMMAND.search(command) or MUTATING_COMMAND.search(command):
            raise AssertionError(f"复验命令{index}包含内部脚本或危险动作")
        expected.append(command)
    parser = _CommandParser()
    parser.feed(html_text)
    parser.close()
    if parser.commands != expected:
        raise AssertionError("JSON与HTML复验命令不一致")
    for command in expected:
        if command not in excel_text:
            raise AssertionError("JSON与Excel复验命令不一致")
    return {"cases": 1, "steps": len(steps), "commands": len(commands)}


__all__ = ["audit_ipsec_artifacts"]
