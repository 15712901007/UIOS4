"""Offline audit for Threat Intelligence Center report artifacts.

The threat-intelligence test is intentionally a single comprehensive pytest
case.  This module validates the three report representations produced by the
project (JSON, HTML and Excel) without connecting to a device.  It checks the
report contract, evidence coverage, and the safety of copy-ready commands.

The audit is deliberately independent of the IOC page objects and backend
implementation.  A failed audit therefore means that the evidence package is
not trustworthy; it never turns a product or environment failure into a pass.
"""

from __future__ import annotations

import json
import re
import shlex
from html.parser import HTMLParser
from pathlib import Path, PureWindowsPath
from typing import Dict, Iterable, List, Mapping, Sequence, Tuple


REQUIRED_SECTIONS: Tuple[str, ...] = (
    "【测试操作】",
    "【页面验证】",
    "【后端验证】",
    "【运行时验证】",
    "【协议验证】",
    "【清理结果】",
)

# The labels are kept as aliases because the UI has used both a short label and
# a label with a "中心" suffix across firmware builds.  The canonical name is
# what callers should put in a report; any alias is accepted for auditability.
REQUIRED_PAGES: Mapping[str, Tuple[str, ...]] = {
    "威胁态势": ("威胁态势", "威胁态势中心"),
    "威胁监控": ("威胁监控",),
    "IOC管理": ("IOC管理", "IOC 管理", "IOC配置", "IOC 配置"),
    "命中告警": ("命中告警", "告警中心", "IOC告警", "IOC 告警"),
    "事件响应": ("事件响应", "响应处置", "处置中心"),
    "报表中心": ("报表中心", "威胁报表", "IOC报表"),
}
REQUIRED_SUBPAGES: Mapping[str, Tuple[str, ...]] = {
    "外界日志": ("外界日志", "外部日志", "系统日志", "Syslog", "syslog"),
    "黑名单": ("黑名单", "IOC黑名单", "IOC 黑名单"),
    "白名单": ("白名单", "IOC白名单", "IOC 白名单"),
}
REQUIRED_SCRIPTS: Tuple[str, ...] = (
    "ioc_overview",
    "ioc_homepage",
    "ioc_monitor",
    "ioc_detail",
    "ioc_syslog",
    "ioc_alert",
    "ioc_policy",
    "ioc_blacklist",
    "ioc_whitelist",
    "ioc_report",
)

INTERNAL_STATUSES = {
    "passed", "failed", "error", "warning", "not_applicable", "skipped",
}

# These patterns are applied to command text, not to the report body.  IOC
# values (including an IP/MAC indicator) are legitimate evidence and must not
# be rejected merely because they appear in a threat-intelligence report.
SHELL_SCRIPT_TOKENS = re.compile(
    r"\b(?:if|then|fi|for|do|done|while|case|esac)\b|"
    r"\$\(|\$\{|\$(?:\?|[A-Za-z_][A-Za-z0-9_]*)|"
    # A shell assignment is only meaningful at command start or after a
    # control operator.  IOC query arguments such as ``TYPE=stats`` and
    # ``time_range=today`` are ordinary positional arguments and must remain
    # valid copy-ready commands.
    r"(?:^|[;&|]\s*)(?:[A-Za-z_][A-Za-z0-9_]*)\s*=|"
    r"__[A-Z0-9_]+__|\bbase64(?:\s|$)|"
    r"(?:^|[ ;])\[(?:router|client|peer)\](?:$|[ ;])|"
    + re.escape("'\"'\"'")
)

MUTATING_COMMAND = re.compile(
    r"(?i)(?:^|[ ;|])(?:rm|mv|cp|kill|pkill|reboot|shutdown|service|systemctl|"
    r"touch|mkdir|rmdir)\s|"
    r"\b(?:delete|insert|update|drop|truncate|replace)\b|"
    r"\b(?:ip|ipset|iptables|ip6tables)\s+(?:addr|route|rule|link|xfrm|"
    r"add|del|delete|flush|destroy|create|restore|append|-a|-d|-f)\b|"
    r"\bvtysh\b.*\bconfigure\b|"
    r"/usr/ikuai/(?:script|function)/[^\s;]+\s+(?:init|add|edit|del|up|down|"
    r"save|put|post)\b|"
    r"\bcurl\b[^\r\n]*(?:\s-X\s*(?:POST|PUT|PATCH|DELETE)|--data(?:-raw)?|"
    r"--upload-file|\s-d\s)"
)

SENSITIVE_ARGUMENT = re.compile(
    r"(?ix)"
    r"(?:\b(?:password|passwd|secret|token|api[_ -]?key|access[_ -]?token|"
    r"auth(?:entication)?(?:[_ -]?(?:pass|password|key|token))?|cookie|"
    r"private[_ -]?key|client[_ -]?secret)\b\s*(?:=|:)\s*"
    r"(?!<[^>]+>|\[已隐藏\]|\[命令已隐藏[^\]]*\]|\{(?:configured|stored|missing)\}|"
    r"(?:configured|stored|missing|omitted|redacted)\b|\*{3,})"
    r"[^\s,;)}\]]+|"
    r"(?:--(?:token|api[-_]?key|password|secret|cookie)|"
    r"-H\s+['\"]?(?:authorization|x-api-key|x-auth-token)\s*:)\s*"
    r"[^\s'\"]+"
    r")"
)

LOCAL_USER_PATH = re.compile(
    r"(?i)(?:[a-z]:[\\/]+users[\\/]+[^\\/\s\"'<>]+|"
    r"file:///(?:[a-z]:)?[\\/]+users[\\/]+[^\s\"'<>]+)"
)


class _VerificationCommandParser(HTMLParser):
    """Collect command code nodes in browser-decoded order."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.commands: List[str] = []
        self._parts: List[str] | None = None

    def handle_starttag(self, tag, attrs) -> None:
        if tag.lower() != "code" or self._parts is not None:
            return
        element_id = str(dict(attrs).get("id", ""))
        if element_id.startswith("verification-command-"):
            self._parts = []

    def handle_data(self, data) -> None:
        if self._parts is not None:
            self._parts.append(data)

    def handle_endtag(self, tag) -> None:
        if tag.lower() == "code" and self._parts is not None:
            self.commands.append("".join(self._parts))
            self._parts = None


def _html_commands(html_text: str) -> List[str]:
    parser = _VerificationCommandParser()
    parser.feed(html_text)
    parser.close()
    return parser.commands


def _is_true(value) -> bool:
    return value is True or str(value).strip().lower() in {
        "true", "yes", "on", "y", "1", "是",
    }


def _is_absolute_path(value: str) -> bool:
    text = str(value or "").strip()
    return bool(text) and (
        Path(text).is_absolute() or PureWindowsPath(text).is_absolute()
    )


def _workbook_data(path: Path) -> Tuple[str, List[str]]:
    """Return all cell text and the reconstructed command sequence.

    ``test_results_to_excel`` splits long command strings across rows.  Using
    the command number and part marker here preserves exact command equality,
    including commands longer than Excel's display width.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is in requirements
        raise AssertionError("审计Excel需要openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        all_values = [
            str(cell.value or "")
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        ]
        if "复验命令" not in workbook.sheetnames:
            raise AssertionError("Excel缺少复验命令sheet")
        sheet = workbook["复验命令"]
        headers = [str(cell.value or "") for cell in sheet[1]]
        required = {"命令序号", "命令", "命令分片"}
        missing = sorted(required - set(headers))
        if missing:
            raise AssertionError("Excel复验命令sheet缺少字段：" + ",".join(missing))
        index = {name: headers.index(name) + 1 for name in required}

        # Key by step number and command number.  The writer emits rows in
        # command order, and command-part values are already in sequence.
        chunks: Dict[Tuple[str, str], List[str]] = {}
        order: List[Tuple[str, str]] = []
        for row_number in range(2, sheet.max_row + 1):
            command_no = sheet.cell(row_number, index["命令序号"]).value
            command_text = sheet.cell(row_number, index["命令"]).value
            part = sheet.cell(row_number, index["命令分片"]).value
            if command_no in (None, "") or command_text in (None, ""):
                continue
            # Step number is column B in the project exporter.  It is not
            # required as a header by older exports, so use the value directly.
            step_no = sheet.cell(row_number, 2).value
            key = (str(step_no or ""), str(command_no))
            if key not in chunks:
                chunks[key] = []
                order.append(key)
            chunks[key].append(str(command_text))
        commands = ["".join(chunks[key]) for key in order]
        return "\n".join(all_values), commands
    finally:
        workbook.close()


def _artifact_texts(
    json_path: Path, html_path: Path, excel_path: Path
) -> Tuple[Dict[str, str], Dict, str, List[str]]:
    json_text = json_path.read_text(encoding="utf-8")
    html_text = html_path.read_text(encoding="utf-8")
    data = json.loads(json_text)
    excel_text, excel_commands = _workbook_data(excel_path)
    return (
        {"json": json_text, "html": html_text, "excel": excel_text},
        data,
        html_text,
        excel_commands,
    )


def _assert_forbidden_values(
    artifacts: Mapping[str, str], sensitive_values: Iterable[str]
) -> None:
    for raw in sensitive_values or ():
        value = str(raw or "")
        if not value.strip() or len(value.strip()) < 4:
            continue
        locations = [name for name, text in artifacts.items() if value in text]
        if locations:
            raise AssertionError(
                "内存登记敏感值命中产物位置：" + ",".join(sorted(locations))
            )
    locations = [
        name for name, text in artifacts.items() if LOCAL_USER_PATH.search(text)
    ]
    if locations:
        raise AssertionError("本机用户路径命中产物位置：" + ",".join(sorted(locations)))


def _assert_no_excel_formula_injection(values: Sequence[str]) -> None:
    for value in values:
        if str(value).lstrip().startswith(("=", "+", "-", "@")):
            raise AssertionError("Excel包含潜在公式注入单元格")


def _normalise_text(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).casefold()


def _coverage_hits(text: str, aliases: Mapping[str, Sequence[str]]) -> Dict[str, str]:
    normalised = _normalise_text(text)
    missing: Dict[str, str] = {}
    for canonical, options in aliases.items():
        if not any(_normalise_text(option) in normalised for option in options):
            missing[canonical] = canonical
    return missing


def _step_coverage_text(steps: Sequence[Mapping]) -> str:
    parts: List[str] = []
    for step in steps:
        parts.append(str(step.get("name", "")))
        parts.append(str(step.get("description", "")))
        parts.extend(str(item) for item in (step.get("details", []) or []))
        for command in (step.get("verification_commands", []) or []):
            if isinstance(command, Mapping):
                parts.extend(str(command.get(key, "")) for key in (
                    "purpose", "expected", "actual", "command",
                ))
            else:
                parts.append(str(command))
    return "\n".join(parts)


def _assert_command_safety(commands: Sequence[Mapping]) -> List[str]:
    expected: List[str] = []
    for index, item in enumerate(commands, 1):
        if not isinstance(item, Mapping):
            raise AssertionError(f"复验命令{index}不是结构化对象")
        command = str(item.get("command", ""))
        target = str(item.get("target", ""))
        if target not in {"router", "client"}:
            raise AssertionError(f"复验命令{index}目标不合规")
        if not command.strip() or "\n" in command or "\r" in command:
            raise AssertionError(f"复验命令{index}不是单行可执行命令")
        if not _is_true(item.get("copy_ready")):
            raise AssertionError(f"复验命令{index}未标记为可直接复制")
        if _is_true(item.get("contains_secret")):
            raise AssertionError(f"复验命令{index}被标记为包含敏感信息")
        if SHELL_SCRIPT_TOKENS.search(command):
            raise AssertionError(f"复验命令{index}含内部脚本或伪命令语法")
        if MUTATING_COMMAND.search(command):
            raise AssertionError(f"复验命令{index}包含危险动作")
        if SENSITIVE_ARGUMENT.search(command):
            raise AssertionError(f"复验命令{index}含敏感参数")
        try:
            shlex.split(command)
        except ValueError as exc:
            raise AssertionError(f"复验命令{index}引号不完整") from exc
        for key in ("purpose", "expected", "actual", "effect", "valid_when"):
            if not str(item.get(key, "")).strip():
                raise AssertionError(f"复验命令{index}缺少中文元数据")
        expected.append(command)
    return expected


def audit_ioc_artifacts(
    json_path,
    html_path,
    excel_path,
    sensitive_values: Iterable[str] = (),
    *,
    required_pages: Mapping[str, Sequence[str]] | None = None,
    required_subpages: Mapping[str, Sequence[str]] | None = None,
    required_scripts: Iterable[str] | None = None,
    require_subpages: bool = True,
) -> Dict[str, object]:
    """Validate one Threat Intelligence Center artifact triplet.

    The function reads only local files.  It returns counts and artifact
    basenames on success and raises :class:`AssertionError` with a safe,
    non-sensitive location/ordinal on contract violations.
    """
    json_path, html_path, excel_path = map(
        Path, (json_path, html_path, excel_path)
    )
    artifacts, data, html_text, excel_commands = _artifact_texts(
        json_path, html_path, excel_path
    )
    json_text = artifacts["json"]
    excel_text = artifacts["excel"]
    _assert_forbidden_values(artifacts, sensitive_values)

    cases = data.get("test_cases", []) or []
    if len(cases) != 1:
        raise AssertionError("威胁情报产物必须且只能包含一个综合用例")
    case = cases[0]
    original_name = str(case.get("original_name", "")).split("[")[0]
    if original_name != "test_threat_intelligence_comprehensive":
        raise AssertionError("JSON不是威胁情报唯一综合用例产物")
    if "威胁情报中心" not in str(case.get("name", "")):
        raise AssertionError("JSON用例中文名不是威胁情报中心")

    case_status = str(case.get("status", "")).strip().lower()
    if case_status not in INTERNAL_STATUSES:
        raise AssertionError("威胁情报用例状态不合法")
    screenshot = str(case.get("screenshot_path", "") or "")
    if _is_absolute_path(screenshot):
        raise AssertionError("截图路径必须是相对路径")

    steps = case.get("steps", []) or []
    if not steps:
        raise AssertionError("威胁情报JSON缺少步骤明细")
    allowed_step_statuses = INTERNAL_STATUSES
    for index, step in enumerate(steps, 1):
        name = str(step.get("name", ""))
        if "操作：" not in name or "；验证：" not in name:
            raise AssertionError(f"步骤{index}标题不符合中文操作/验证格式")
        status = str(step.get("status", "")).strip().lower()
        if status not in allowed_step_statuses:
            raise AssertionError(f"步骤{index}状态不合法")
        details = "\n".join(map(str, step.get("details", []) or []))
        missing = [section for section in REQUIRED_SECTIONS if section not in details]
        if missing:
            raise AssertionError(f"步骤{index}缺少六段证据：{len(missing)}段")

    # Validate top-level counts when present, but stay compatible with older
    # hand-authored archives which predate total_steps.
    if "total" in data and int(data.get("total") or 0) != len(cases):
        raise AssertionError("JSON用例总数与明细不一致")
    if "total_steps" in data and int(data.get("total_steps") or 0) != len(steps):
        raise AssertionError("JSON步骤总数与明细不一致")

    step_text = _step_coverage_text(steps)
    pages = required_pages or REQUIRED_PAGES
    missing_pages = _coverage_hits(step_text, pages)
    if missing_pages:
        raise AssertionError("缺少页面覆盖：" + ",".join(missing_pages))
    if require_subpages:
        subpages = required_subpages or REQUIRED_SUBPAGES
        missing_subpages = _coverage_hits(step_text, subpages)
        if missing_subpages:
            raise AssertionError("缺少子页面覆盖：" + ",".join(missing_subpages))

    scripts = tuple(required_scripts or REQUIRED_SCRIPTS)
    lower_text = step_text.casefold()
    missing_scripts = [
        script for script in scripts if str(script).casefold() not in lower_text
    ]
    if missing_scripts:
        raise AssertionError("缺少底层脚本覆盖：" + ",".join(missing_scripts))

    commands: List[Mapping] = []
    for step in steps:
        raw = step.get("verification_commands", []) or []
        if isinstance(raw, (str, bytes, Mapping)):
            raw = [raw]
        commands.extend(raw)
    if not commands:
        raise AssertionError("威胁情报产物缺少人工复验命令")
    expected_commands = _assert_command_safety(commands)

    html_commands = _html_commands(html_text)
    if html_commands != expected_commands:
        raise AssertionError("复验命令在JSON/HTML中数量、内容或顺序不一致")
    if excel_commands != expected_commands:
        raise AssertionError("复验命令在JSON/Excel中数量、内容或顺序不一致")

    # Formula injection checks use every workbook cell, not just the command
    # column.  Re-read the workbook values through the already-built text is
    # insufficient for distinguishing a leading formula from ordinary text;
    # the command-column reconstruction above remains the exact consistency
    # source, while this conservative scan protects all visible cells.
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise AssertionError("审计Excel需要openpyxl") from exc
    workbook = load_workbook(excel_path, read_only=True, data_only=False)
    try:
        workbook_values = [
            str(cell.value or "")
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        ]
    finally:
        workbook.close()
    _assert_no_excel_formula_injection(workbook_values)

    return {
        "cases": 1,
        "steps": len(steps),
        "commands": len(commands),
        "pages": len(pages),
        "subpages": len(required_subpages or REQUIRED_SUBPAGES) if require_subpages else 0,
        "scripts": len(scripts),
        "artifacts": {
            "json": json_path.name,
            "html": html_path.name,
            "excel": excel_path.name,
        },
    }


__all__ = [
    "REQUIRED_SECTIONS",
    "REQUIRED_PAGES",
    "REQUIRED_SUBPAGES",
    "REQUIRED_SCRIPTS",
    "audit_ioc_artifacts",
]
