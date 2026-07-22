"""Validate archived SNMP JSON/HTML/Excel reports without printing secrets."""

from __future__ import annotations

import json
import re
from html.parser import HTMLParser
from pathlib import Path
from typing import Dict, List


SHELL_SCRIPT_TOKENS = re.compile(
    r"\b(?:if|then|fi|for|do|done|while|case|esac)\b|"
    r"\$\(|\$\{|\$(?:\?|[A-Za-z_][A-Za-z0-9_]*)|"
    r"(?:^|[ ;])rc\s*=|__[A-Z0-9_]+__|\bbase64(?:\s|$)|"
    r"(?:^|[ ;])\[(?:router|client)\](?:$|[ ;])|"
    + re.escape("'\"'\"'")
)
REQUIRED_SECTIONS = (
    "【测试操作】", "【页面验证】", "【后端验证】",
    "【运行时验证】", "【协议验证】", "【清理结果】",
)
SNMP_SECRET_ARGUMENTS = re.compile(
    r"(?i:\bsnmp(?:get|walk|bulkwalk)\b)[^\r\n]*(?:"
    r"\s-(?:c|A|X)(?:\s+|=)?[^\s]+|"
    r"\s--(?i:community|auth-pass|priv-pass)(?:\s+|=)[^\s]+)"
)
SENSITIVE_ASSIGNMENTS = re.compile(
    r"(?i)\b(?:community|auth_pass|priv_pass|password)\s*="
)


class _VerificationCommandParser(HTMLParser):
    """Collect verification command text exactly as a browser decodes it."""

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.commands: List[str] = []
        self._parts: List[str] | None = None

    def handle_starttag(self, tag, attrs):
        if tag.lower() != "code" or self._parts is not None:
            return
        attributes = dict(attrs)
        element_id = str(attributes.get("id", ""))
        if element_id.startswith("verification-command-"):
            self._parts = []

    def handle_data(self, data):
        if self._parts is not None:
            self._parts.append(data)

    def handle_endtag(self, tag):
        if tag.lower() == "code" and self._parts is not None:
            self.commands.append("".join(self._parts))
            self._parts = None


def _html_verification_commands(html_text: str) -> List[str]:
    parser = _VerificationCommandParser()
    parser.feed(html_text)
    parser.close()
    return parser.commands


def _is_true(value) -> bool:
    return value is True or str(value).strip().lower() in {
        "true", "yes", "on", "y", "1", "是",
    }


def _workbook_text(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return "\n".join(
            str(cell.value or "")
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
    finally:
        workbook.close()


def audit_snmp_artifacts(json_path, html_path, excel_path) -> Dict[str, int]:
    json_path = Path(json_path)
    html_path = Path(html_path)
    excel_path = Path(excel_path)
    data = json.loads(json_path.read_text(encoding="utf-8"))
    html_text = html_path.read_text(encoding="utf-8")
    excel_text = _workbook_text(excel_path)
    cases = data.get("test_cases", [])
    if len(cases) != 1 or "SNMP" not in str(cases[0].get("name", "")):
        raise AssertionError("JSON不是唯一SNMP综合用例产物")
    steps = cases[0].get("steps", []) or []
    if not steps:
        raise AssertionError("SNMP JSON缺少步骤明细")
    for index, step in enumerate(steps, 1):
        name = str(step.get("name", ""))
        if "操作：" not in name or "；验证：" not in name:
            raise AssertionError(f"步骤{index}标题不符合中文操作/验证格式")
        details = "\n".join(map(str, step.get("details", []) or []))
        missing = [section for section in REQUIRED_SECTIONS if section not in details]
        if missing:
            raise AssertionError(f"步骤{index}缺少六段证据：{len(missing)}段")

    commands = [
        command
        for step in steps
        for command in (step.get("verification_commands", []) or [])
    ]
    for index, item in enumerate(commands, 1):
        command = str(item.get("command", ""))
        if item.get("target") not in {"router", "client"}:
            raise AssertionError(f"复验命令{index}目标不合规")
        if not command or "\n" in command or "\r" in command:
            raise AssertionError(f"复验命令{index}不是单行可执行命令")
        if _is_true(item.get("contains_secret")):
            raise AssertionError(f"复验命令{index}被标记为包含敏感信息")
        if not _is_true(item.get("copy_ready")):
            raise AssertionError(f"复验命令{index}未标记为可直接复制")
        if SNMP_SECRET_ARGUMENTS.search(command) or SENSITIVE_ASSIGNMENTS.search(command):
            raise AssertionError(f"复验命令{index}含SNMP敏感参数")
        if SHELL_SCRIPT_TOKENS.search(command):
            raise AssertionError(f"复验命令{index}含内部脚本或伪命令语法")
        if any(not str(item.get(key, "")).strip() for key in (
            "purpose", "expected", "actual", "effect", "valid_when"
        )):
            raise AssertionError(f"复验命令{index}缺少中文元数据")

    expected_commands = [str(command.get("command", "")) for command in commands]
    html_commands = _html_verification_commands(html_text)
    if len(html_commands) != len(expected_commands):
        raise AssertionError(
            "复验命令在JSON/HTML中数量不一致: "
            f"JSON={len(expected_commands)}, HTML={len(html_commands)}"
        )
    for index, (expected, actual) in enumerate(
        zip(expected_commands, html_commands), 1
    ):
        if actual != expected:
            raise AssertionError(
                f"复验命令{index}在JSON/HTML中内容或顺序不一致"
            )
        if expected not in excel_text:
            raise AssertionError(f"复验命令{index}在JSON/Excel中不一致")
    if "完整源码堆栈已按凭据安全策略隐藏" not in (
        str(cases[0].get("error_traceback", "")) + " " + html_text
    ) and cases[0].get("status") == "failed":
        raise AssertionError("失败报告未隐藏源码堆栈")
    return {"cases": 1, "steps": len(steps), "commands": len(commands)}


__all__ = ["audit_snmp_artifacts"]
