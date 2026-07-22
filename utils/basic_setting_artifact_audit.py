"""安全审计归档后的基础设置 JSON、HTML 与 Excel 产物。

审计失败只报告字段位置或命令序号，不回显命令正文、敏感值或本机绝对
路径。成功结果仅包含数量和三个产物的 basename。
"""

from __future__ import annotations

import json
import re
from html.parser import HTMLParser
from pathlib import Path, PureWindowsPath
from typing import Dict, Iterable, List


REQUIRED_SECTIONS = (
    "【测试操作】",
    "【页面验证】",
    "【后端验证】",
    "【运行时验证】",
    "【协议验证】",
    "【清理结果】",
)
INTERNAL_STATUSES = {
    "passed", "failed", "error", "warning", "not_applicable", "skipped",
}
STATUS_CN = {
    "passed": "通过",
    "failed": "失败",
    "error": "失败",
    "warning": "警告",
    "not_applicable": "不适用",
    "skipped": "不适用",
}
SHELL_SCRIPT_TOKENS = re.compile(
    r"\b(?:if|then|fi|for|do|done|while|case|esac)\b|"
    r"\$\(|\$\{|\$(?:\?|[A-Za-z_][A-Za-z0-9_]*)|"
    r"(?:^|[ ;])rc\s*=|__[A-Z0-9_]+__|\bbase64(?:\s|$)|"
    r"(?:^|[ ;])\[(?:router|client)\](?:$|[ ;])|"
    r"(?:^|[ ;])set\s+-[a-z]+|"
    + re.escape("'\"'\"'")
)
SENSITIVE_COMMAND_TOKENS = re.compile(
    r"(?i)\b(?:password|passwd|secret|community|auth_pass|priv_pass|"
    r"console_password)\b\s*(?:=|:|\s)"
)
MUTATING_RECOVERY_COMMANDS = re.compile(
    r"(?i)(?:UPDATE\s+basic\b|/usr/ikuai/script/basic\.sh\s+init\b|"
    r"\bip\s+route\s+(?:replace|del)\b|\bpkill\b|"
    r"\brm\s+-f\s+/tmp/ikuai-basic-)"
)
HARDWARE_ADDRESS = re.compile(
    r"(?i)(?<![0-9a-f])(?:[0-9a-f]{2}[:-]){5}[0-9a-f]{2}(?![0-9a-f])"
)
LOCAL_USER_PATH = re.compile(
    r"(?i)(?:[a-z]:[\\/]+users[\\/]+[^\\/\s\"'<>]+|"
    r"/home/[^/\s\"'<>]+)"
)


class _CommandParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.commands: List[str] = []
        self._parts: List[str] | None = None

    def handle_starttag(self, tag, attrs):
        if tag.lower() != "code" or self._parts is not None:
            return
        element_id = str(dict(attrs).get("id", ""))
        if element_id.startswith("verification-command-"):
            self._parts = []

    def handle_data(self, data):
        if self._parts is not None:
            self._parts.append(data)

    def handle_endtag(self, tag):
        if tag.lower() == "code" and self._parts is not None:
            self.commands.append("".join(self._parts))
            self._parts = None


def _html_commands(html_text: str) -> List[str]:
    parser = _CommandParser()
    parser.feed(html_text)
    parser.close()
    return parser.commands


def _is_true(value) -> bool:
    return value is True or str(value).strip().lower() in {
        "true", "yes", "on", "y", "1", "是",
    }


def _excel_data(path: Path):
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        text = "\n".join(
            str(cell.value or "")
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        if "复验命令" not in workbook.sheetnames:
            raise AssertionError("Excel缺少复验命令sheet")
        sheet = workbook["复验命令"]
        headers = [str(cell.value or "") for cell in sheet[1]]
        if "命令" not in headers:
            raise AssertionError("Excel复验命令sheet缺少命令列")
        command_column = headers.index("命令") + 1
        commands = [
            str(sheet.cell(row=row, column=command_column).value)
            for row in range(2, sheet.max_row + 1)
            if sheet.cell(row=row, column=command_column).value not in (None, "")
        ]
        return text, commands
    finally:
        workbook.close()


def _is_absolute_path(value: str) -> bool:
    text = str(value or "").strip()
    return bool(text) and (
        Path(text).is_absolute() or PureWindowsPath(text).is_absolute()
    )


def _assert_no_forbidden_values(
    artifacts: Dict[str, str], forbidden_values: Iterable[str]
):
    for value in forbidden_values or ():
        secret = str(value or "")
        if not secret or not secret.strip():
            continue
        locations = [name for name, text in artifacts.items() if secret in text]
        if locations:
            raise AssertionError(
                "内存登记敏感值命中产物位置：" + ",".join(sorted(locations))
            )


def _assert_no_pattern_leaks(artifacts: Dict[str, str]):
    checks = (
        (HARDWARE_ADDRESS, "硬件地址"),
        (LOCAL_USER_PATH, "本机用户路径"),
    )
    for pattern, label in checks:
        locations = [
            name for name, content in artifacts.items() if pattern.search(content)
        ]
        if locations:
            raise AssertionError(
                f"{label}命中产物位置：" + ",".join(sorted(locations))
            )


def audit_basic_setting_artifacts(
    json_path,
    html_path,
    excel_path,
    *,
    forbidden_values: Iterable[str] = (),
) -> Dict[str, object]:
    """Validate one archived basic-setting report set without echoing content."""
    json_path = Path(json_path)
    html_path = Path(html_path)
    excel_path = Path(excel_path)
    json_text = json_path.read_text(encoding="utf-8")
    html_text = html_path.read_text(encoding="utf-8")
    data = json.loads(json_text)
    excel_text, excel_commands = _excel_data(excel_path)

    artifact_text = {
        "json": json_text,
        "html": html_text,
        "excel": excel_text,
    }
    _assert_no_forbidden_values(artifact_text, forbidden_values)
    _assert_no_pattern_leaks(artifact_text)

    cases = data.get("test_cases", []) or []
    if len(cases) != 1:
        raise AssertionError("基础设置产物必须且只能包含一个综合用例")
    case = cases[0]
    original_name = str(case.get("original_name", "")).split("[")[0]
    if original_name != "test_basic_setting_comprehensive":
        raise AssertionError("JSON不是基础设置唯一综合用例产物")
    case_status = str(case.get("status", "")).strip().lower()
    if case_status not in INTERNAL_STATUSES:
        raise AssertionError("基础设置用例状态不属于四态")
    screenshot = str(case.get("screenshot_path", "") or "")
    if _is_absolute_path(screenshot):
        raise AssertionError("基础设置截图路径必须是相对路径")

    steps = case.get("steps", []) or []
    if not steps:
        raise AssertionError("基础设置JSON缺少步骤明细")
    for step_index, step in enumerate(steps, 1):
        name = str(step.get("name", ""))
        if "操作：" not in name or "；验证：" not in name:
            raise AssertionError(
                f"步骤{step_index}标题不符合中文操作/验证格式"
            )
        status = str(step.get("status", "")).strip().lower()
        if status not in INTERNAL_STATUSES:
            raise AssertionError(f"步骤{step_index}状态不属于四态")
        details = "\n".join(map(str, step.get("details", []) or []))
        missing = [section for section in REQUIRED_SECTIONS if section not in details]
        if missing:
            raise AssertionError(
                f"步骤{step_index}缺少六段证据：{len(missing)}段"
            )

    commands = [
        item
        for step in steps
        for item in (step.get("verification_commands", []) or [])
    ]
    if not commands:
        raise AssertionError("基础设置产物缺少人工复验命令")
    for command_index, item in enumerate(commands, 1):
        command = str(item.get("command", ""))
        if item.get("target") not in {"router", "client"}:
            raise AssertionError(f"复验命令{command_index}目标不合规")
        if not command or "\n" in command or "\r" in command:
            raise AssertionError(f"复验命令{command_index}不是单行可执行命令")
        if not _is_true(item.get("copy_ready")):
            raise AssertionError(f"复验命令{command_index}未标记为可直接复制")
        if _is_true(item.get("contains_secret")):
            raise AssertionError(f"复验命令{command_index}被标记为包含敏感信息")
        if any(not str(item.get(key, "")).strip() for key in (
            "purpose", "expected", "actual", "effect", "valid_when",
        )):
            raise AssertionError(f"复验命令{command_index}缺少中文元数据")
        if SHELL_SCRIPT_TOKENS.search(command):
            raise AssertionError(f"复验命令{command_index}含内部脚本语法")
        if SENSITIVE_COMMAND_TOKENS.search(command):
            raise AssertionError(f"复验命令{command_index}含凭据参数")
        if MUTATING_RECOVERY_COMMANDS.search(command):
            raise AssertionError(f"复验命令{command_index}暴露恢复或清理动作")

    expected_commands = [str(item.get("command", "")) for item in commands]
    html_commands = _html_commands(html_text)
    if html_commands != expected_commands:
        raise AssertionError("复验命令在JSON/HTML中数量、内容或顺序不一致")
    if excel_commands != expected_commands:
        raise AssertionError("复验命令在JSON/Excel中数量、内容或顺序不一致")

    # Historical ``skipped`` remains an internal input only.
    if "跳过" in html_text or "跳过" in excel_text:
        raise AssertionError("HTML或Excel仍对外显示历史跳过状态")
    expected_labels = {
        STATUS_CN[str(case.get("status", "")).strip().lower()],
        *{
            STATUS_CN[str(step.get("status", "")).strip().lower()]
            for step in steps
        },
    }
    if any(label not in html_text or label not in excel_text for label in expected_labels):
        raise AssertionError("四态中文在HTML或Excel中显示不一致")

    return {
        "cases": 1,
        "steps": len(steps),
        "commands": len(commands),
        "artifacts": {
            "json": json_path.name,
            "html": html_path.name,
            "excel": excel_path.name,
        },
    }


__all__ = ["audit_basic_setting_artifacts"]
