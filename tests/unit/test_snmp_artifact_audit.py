from __future__ import annotations

import json
from copy import deepcopy
from html import escape

import pytest
from openpyxl import load_workbook

from utils.snmp_artifact_audit import audit_snmp_artifacts
from utils.test_results_to_excel import export_results_to_excel


REQUIRED_SECTIONS = (
    "【测试操作】",
    "【页面验证】",
    "【后端验证】",
    "【运行时验证】",
    "【协议验证】",
    "【清理结果】",
)


def _command(index: int = 1, text: str | None = None) -> dict:
    return {
        "target": "router" if index % 2 else "client",
        "target_label": "路由器" if index % 2 else "客户端",
        "host": "192.0.2.1" if index % 2 else "192.0.2.2",
        "shell": "sh",
        "purpose": f"验证SNMP审计命令{index}",
        "command": text or f"grep -F 'snmp-audit-{index}' /var/run/snmp/meta",
        "expected": "输出唯一匹配行",
        "actual": "已获得唯一匹配行",
        "copy_ready": True,
        "effect": "read_only",
        "contains_secret": False,
        "interactive": False,
        "interactive_hint": "无需交互",
        "valid_when": "报告生成后执行",
    }


def _data(commands: list[dict], details: list[str] | None = None) -> dict:
    return {
        "schema_version": 2,
        "total": 1,
        "passed": 1,
        "failed": 0,
        "skipped": 0,
        "total_steps": 1,
        "test_cases": [
            {
                "name": "高级服务-本地服务-SNMP服务",
                "original_name": "test_snmp_server_comprehensive",
                "status": "passed",
                "duration": "1.00s",
                "error_message": None,
                "error_traceback": None,
                "steps": [
                    {
                        "name": "步骤1 操作：执行审计；验证：产物内容一致",
                        "description": "验证JSON、HTML和Excel",
                        "status": "passed",
                        "duration": "1.00s",
                        "details": details or [
                            f"{section}\n不适用：本步骤使用审计夹具"
                            for section in REQUIRED_SECTIONS
                        ],
                        "verification_commands": commands,
                        "actual": {"command_count": len(commands)},
                        "error_message": None,
                    }
                ],
                "step_count": 1,
                "screenshot_path": "",
            }
        ],
    }


def _write_artifacts(tmp_path, data: dict, html_commands=None):
    json_path = tmp_path / "snmp.json"
    html_path = tmp_path / "snmp.html"
    excel_path = tmp_path / "snmp.xlsx"
    json_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    commands = [
        str(item.get("command", ""))
        for step in data["test_cases"][0]["steps"]
        for item in step.get("verification_commands", [])
    ]
    rendered_commands = commands if html_commands is None else list(html_commands)
    code_nodes = [
        "<code id=\"verification-command-0-0-{}\">{}</code>".format(
            index, escape(command, quote=True)
        )
        for index, command in enumerate(rendered_commands, 1)
    ]
    html_path.write_text(
        "<html><body><code id=\"unrelated-example\">忽略</code>"
        + "".join(code_nodes)
        + "</body></html>",
        encoding="utf-8",
    )

    success, message = export_results_to_excel(str(json_path), str(excel_path))
    assert success, message
    return json_path, html_path, excel_path


def test_audit_decodes_html_entities_and_handles_429_commands(tmp_path):
    commands = []
    for index in range(1, 430):
        entity_text = (
            f"entity-{index:03d} <node>&\"quoted\"</node> " + "x" * 180
        )
        commands.append(
            _command(
                index,
                f"grep -F '{entity_text}' /var/run/snmp/snmpd.conf",
            )
        )
    paths = _write_artifacts(tmp_path, _data(commands))

    result = audit_snmp_artifacts(*paths)

    assert result == {"cases": 1, "steps": 1, "commands": 429}


@pytest.mark.parametrize(
    "unsafe_command,contains_secret",
    [
        ("snmpget -v2c -c OMITTED 192.0.2.1 1.3.6.1.2.1.1.5.0", False),
        ("snmpwalk -v3 -u audit -A OMITTED 192.0.2.1 1.3.6.1.2.1.1", False),
        ("snmpget -v3 -u audit -X=OMITTED 192.0.2.1 1.3.6.1.2.1.1.5.0", False),
        ("snmpwalk --community OMITTED 192.0.2.1 1.3.6.1.2.1.1", False),
        ("printf 'community=OMITTED'", False),
        ("grep -F 'safe-snmp-marker' /var/run/snmp/meta", True),
    ],
)
def test_audit_rejects_sensitive_parameters(
    tmp_path, unsafe_command, contains_secret
):
    command = _command(text=unsafe_command)
    command["contains_secret"] = contains_secret
    paths = _write_artifacts(tmp_path, _data([command]))

    with pytest.raises(AssertionError, match="敏感"):
        audit_snmp_artifacts(*paths)


@pytest.mark.parametrize("missing_section", REQUIRED_SECTIONS)
def test_audit_requires_every_evidence_section(tmp_path, missing_section):
    details = [
        f"{section}\n不适用：本步骤使用审计夹具"
        for section in REQUIRED_SECTIONS
        if section != missing_section
    ]
    paths = _write_artifacts(tmp_path, _data([_command()], details=details))

    with pytest.raises(AssertionError, match="缺少六段证据"):
        audit_snmp_artifacts(*paths)


@pytest.mark.parametrize("invalid_target", ["", "server", "router ", "[router]"])
def test_audit_rejects_invalid_target(tmp_path, invalid_target):
    command = _command()
    command["target"] = invalid_target
    paths = _write_artifacts(tmp_path, _data([command]))

    with pytest.raises(AssertionError, match="目标不合规"):
        audit_snmp_artifacts(*paths)


@pytest.mark.parametrize(
    "missing_key", ["purpose", "expected", "actual", "effect", "valid_when"]
)
def test_audit_requires_command_metadata(tmp_path, missing_key):
    command = _command()
    command[missing_key] = "   "
    paths = _write_artifacts(tmp_path, _data([command]))

    with pytest.raises(AssertionError, match="缺少中文元数据"):
        audit_snmp_artifacts(*paths)


def test_audit_rejects_non_copy_ready_command(tmp_path):
    command = _command()
    command["copy_ready"] = False
    paths = _write_artifacts(tmp_path, _data([command]))

    with pytest.raises(AssertionError, match="未标记为可直接复制"):
        audit_snmp_artifacts(*paths)


def test_audit_rejects_html_command_order_mismatch(tmp_path):
    commands = [_command(1), _command(2)]
    html_commands = [commands[1]["command"], commands[0]["command"]]
    paths = _write_artifacts(
        tmp_path, _data(commands), html_commands=html_commands
    )

    with pytest.raises(AssertionError, match="内容或顺序不一致"):
        audit_snmp_artifacts(*paths)


def test_audit_rejects_excel_command_mismatch(tmp_path):
    data = _data([_command()])
    paths = _write_artifacts(tmp_path, data)
    workbook = load_workbook(paths[2])
    try:
        workbook["复验命令"].cell(2, 14, "grep -F 'different' /var/run/snmp/meta")
        workbook.save(paths[2])
    finally:
        workbook.close()

    with pytest.raises(AssertionError, match="JSON/Excel中不一致"):
        audit_snmp_artifacts(*paths)


def test_audit_input_data_is_not_mutated(tmp_path):
    data = _data([_command()])
    original = deepcopy(data)
    paths = _write_artifacts(tmp_path, data)

    audit_snmp_artifacts(*paths)

    assert data == original
