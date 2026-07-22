"""本地服务人工复验命令与报告链路的离线回归测试。"""

from __future__ import annotations

import html as html_lib
import json
import re
from types import SimpleNamespace

import pytest

from utils.replay_commands import build_verification_commands
from utils.report_generator import ReportGenerator
from utils.step_recorder import StepRecorder
from utils.test_results_to_excel import export_results_to_excel
from utils.verify_helper import attach_cmd_recording_to_closure


ROUTER_HOST = "10.66.0.150"
CLIENT_HOST = "10.66.0.18"
SHELL_SCRIPT_TOKENS = re.compile(
    r"\b(?:if|then|fi|for|do|done|while|case|esac)\b|"
    r"\$\(|\$\{|\$(?:\?|[A-Za-z_][A-Za-z0-9_]*)|"
    r"(?:^|[ ;])rc\s*=|__[A-Z0-9_]+__"
)


@pytest.fixture
def backend_stub():
    """只提供命令展示所需的主机配置，不建立任何 SSH 连接。"""
    ssh_config = SimpleNamespace(
        router=SimpleNamespace(host=ROUTER_HOST),
        client=SimpleNamespace(host=CLIENT_HOST),
    )
    return SimpleNamespace(_ssh_config=ssh_config)


# 这些 stub 的名称和签名模拟 BackendVerifier 的公开方法。命令生成器只能
# 检查函数元数据并绑定参数，不能真正调用它们。
def verify_samba_user_database(username, expected=None, must_exist=True):
    raise AssertionError("离线命令生成不应执行 verifier")


def verify_samba_runtime_consistency(prefix=None, expected_firewall_members=None):
    raise AssertionError("离线命令生成不应执行 verifier")


def run_samba_probe(
    username=None,
    password=None,
    host="192.168.148.1",
    iface="ens11",
    operation="list",
    share_name=None,
    control_password=None,
    remote_name=None,
    control_host=None,
    control_iface="ens11",
):
    raise AssertionError("离线命令生成不应执行 verifier")


def verify_ftp_user_database(username, expected_fields=None, must_exist=True):
    raise AssertionError("离线命令生成不应执行 verifier")


def run_ftp_probe(
    username,
    password,
    port,
    host="192.168.148.1",
    iface="ens11",
    operation="list",
    remote_name=None,
    control_password=None,
    control_host="192.168.148.1",
    control_iface="ens11",
    cleanup_username=None,
    cleanup_password=None,
):
    raise AssertionError("离线命令生成不应执行 verifier")


def verify_http_rule_database(
    tagname=None, rule_id=None, expected_fields=None, must_exist=True
):
    raise AssertionError("离线命令生成不应执行 verifier")


def run_http_probe(
    port,
    operation="fetch",
    host="192.168.148.1",
    iface="ens11",
    path="/payload.bin",
    scheme="http",
    server_name=None,
    expected_sha256=None,
    expected_status=None,
    expected_contains=None,
    control_port=None,
    control_host=None,
    control_iface="ens11",
    control_path=None,
    control_scheme=None,
    control_server_name=None,
    rate_limit_kbps=None,
    timeout_seconds=35,
):
    raise AssertionError("离线命令生成不应执行 verifier")


def verify_snmp_database(expected_fields=None, expected_secrets=None):
    raise AssertionError("离线命令生成不应执行 verifier")


def verify_snmp_singleton_contract():
    raise AssertionError("离线命令生成不应执行 verifier")


def verify_snmp_generated_config(
    expected_fields=None, expected_secrets=None, expect_present=True
):
    raise AssertionError("离线命令生成不应执行 verifier")


def verify_snmp_runtime_consistency(expected_fields=None, expected_secrets=None):
    raise AssertionError("离线命令生成不应执行 verifier")


def verify_snmp_test_artifacts_absent(prefix, candidate_ports, snapshot=None):
    raise AssertionError("离线命令生成不应执行 verifier")


def verify_snmp_client_route(host, expected_iface):
    raise AssertionError("离线命令生成不应执行 verifier")


def run_snmp_probe(
    version,
    host,
    oid,
    operation="get",
    community=None,
    username=None,
    security=None,
    auth_proto=None,
    auth_pass=None,
    priv_proto=None,
    priv_pass=None,
    expect_success=True,
    expected_value=None,
    expected_failure=None,
):
    raise AssertionError("离线命令生成不应执行 verifier")


def _commands(backend_stub, verifier, *args, **kwargs):
    commands = build_verification_commands(
        backend_stub,
        verifier,
        args=args,
        kwargs=kwargs,
    )
    assert commands is not None
    return commands


def _assert_copy_ready(command_record):
    command = command_record["command"]
    assert command == command.strip()
    assert "\n" not in command and "\r" not in command
    assert command_record["copy_ready"] is True
    assert command_record["effect"] == "read_only"
    assert command_record["contains_secret"] is False
    assert command_record["host"] in {ROUTER_HOST, CLIENT_HOST}
    assert command_record["target_label"] in {"路由器", "测试客户端"}

    # 目标标签和自动化内部脚本语法不得混入可复制文本。
    assert "[router]" not in command
    assert "[client]" not in command
    assert "<redacted>" not in command
    assert "'\"'\"'" not in command
    assert SHELL_SCRIPT_TOKENS.search(command) is None


def test_samba_user_sql_is_complete_copy_ready_and_not_truncated(backend_stub):
    username = "smb_t_038d_ro"
    command_record = _commands(
        backend_stub,
        verify_samba_user_database,
        username,
        expected={"perm": "ro", "guest": "yes", "browseable": "no"},
    )[0]

    _assert_copy_ready(command_record)
    command = command_record["command"]
    assert command_record["target"] == "router"
    assert command_record["host"] == ROUTER_HOST
    assert len(command) > 200
    assert "CASE WHEN length(COALESCE(passwd,''))>0" in command
    assert f"WHERE username='{username}' LIMIT 1;" in command
    assert command.endswith('LIMIT 1;"')
    assert "…" not in command and "..." not in command


def test_samba_runtime_commands_are_independent_and_human_runnable(backend_stub):
    commands = _commands(
        backend_stub,
        verify_samba_runtime_consistency,
        "smb_t_038d_",
    )

    assert len(commands) >= 10
    for command_record in commands:
        _assert_copy_ready(command_record)
        assert command_record["target"] == "router"

    command_texts = [item["command"] for item in commands]
    assert any("FROM smbd_dir" in command for command in command_texts)
    assert "pidof ik_smbd" in command_texts
    assert "pidof nmbd" in command_texts
    assert "pidof wsdd2" in command_texts
    assert "ipset list DROP_T_PORTS_WAN_IN" in command_texts
    assert "ipset list DROP_U_PORTS_WAN_IN" in command_texts


def test_ftp_and_http_representative_commands_use_the_correct_target(backend_stub):
    ftp_db = _commands(
        backend_stub,
        verify_ftp_user_database,
        "ftp_t_demo_ro",
        expected_fields={"permission": "ro"},
    )[0]
    ftp_probe_commands = _commands(
        backend_stub,
        run_ftp_probe,
        "ftp_t_demo_ro",
        "ftp-password-not-for-report",
        2121,
        operation="list",
    )
    ftp_probe = next(
        item for item in ftp_probe_commands if "curl " in item["command"]
    )
    http_db = _commands(
        backend_stub,
        verify_http_rule_database,
        "http_t_demo",
        expected_fields={"http_port": 18080},
    )[0]
    http_probe = _commands(
        backend_stub,
        run_http_probe,
        18080,
        server_name="files.test",
        expected_sha256="a" * 64,
    )

    for command_record in [ftp_db, *ftp_probe_commands, http_db, *http_probe]:
        _assert_copy_ready(command_record)

    assert ftp_db["target"] == "router"
    assert "FROM ftp_server" in ftp_db["command"]
    assert ftp_probe["target"] == "client"
    assert ftp_probe["host"] == CLIENT_HOST
    assert ftp_probe["interactive"] is True
    assert '--user "ftp_t_demo_ro"' in ftp_probe["command"]
    assert "ftp://192.168.148.1:2121/" in ftp_probe["command"]

    assert http_db["target"] == "router"
    assert "FROM http_server" in http_db["command"]
    assert all(item["target"] == "client" for item in http_probe)
    assert any("--write-out" in item["command"] for item in http_probe)
    assert any(item["command"].endswith("| sha256sum") for item in http_probe)


def test_ftp_and_samba_passwords_never_enter_report_payload(backend_stub):
    secrets = {
        "ftp-password-9Kq!",
        "ftp-control-password-2Vr!",
        "ftp-cleanup-password-7Ws!",
        "samba-password-8Yx!",
        "samba-control-password-4Tm!",
    }
    ftp_commands = _commands(
        backend_stub,
        run_ftp_probe,
        "ftp_t_secret_user",
        "ftp-password-9Kq!",
        2121,
        control_password="ftp-control-password-2Vr!",
        cleanup_password="ftp-cleanup-password-7Ws!",
    )
    samba_commands = _commands(
        backend_stub,
        run_samba_probe,
        username="smb_t_secret_user",
        password="samba-password-8Yx!",
        share_name="smb_t_secret_share",
        control_password="samba-control-password-4Tm!",
    )

    report_payload = json.dumps(
        ftp_commands + samba_commands,
        ensure_ascii=False,
        sort_keys=True,
    )
    for secret in secrets:
        assert secret not in report_payload
    assert "<redacted>" not in report_payload
    assert any(item["interactive"] is True for item in ftp_commands)
    assert any(item["interactive"] is True for item in samba_commands)


def test_step_recorder_serializes_verification_commands_without_command_loss(
    backend_stub,
):
    source = _commands(
        backend_stub,
        verify_samba_user_database,
        "smb_t_serialize_ro",
    )[0]
    recorder = StepRecorder()

    with recorder.step("添加RO匿名隐藏共享用户", "校验DB和运行时映射"):
        recorder.add_detail("【页面验证】✓ 用户已出现在列表中")
        recorder.add_verification_command(source)

    steps = recorder.get_steps()
    serialized = json.loads(json.dumps(steps, ensure_ascii=False))
    stored = serialized[0]["verification_commands"][0]

    assert serialized[0]["name"] == "添加RO匿名隐藏共享用户"
    assert stored["target_label"] == "路由器"
    assert stored["host"] == ROUTER_HOST
    assert stored["target"] == "router"
    assert stored["shell"] == "sh"
    assert stored["purpose"] == source["purpose"]
    assert stored["expected"] == source["expected"]
    assert stored["effect"] == "read_only"
    assert stored["copy_ready"] is True
    assert stored["contains_secret"] is False
    assert stored["interactive"] is False
    assert stored["valid_when"] == source["valid_when"]
    assert stored["command"] == source["command"]
    assert len(stored["command"]) > 200


def test_local_service_wrapper_hides_internal_script_and_records_manual_command(
    backend_stub,
):
    secret = "wrapper-secret-must-never-appear"
    backend_stub.mark_cmd_start = lambda: (0, 0)
    backend_stub.collect_cmds_since_mark = lambda mark: [
        f"[client] if test -n '{secret}'; then ftp_probe; fi"
    ]

    def ssh_verify(label, verify_func, *args, must_pass=False, **kwargs):
        return SimpleNamespace(passed=True, details={}, raw_output="")

    def ftp_probe_with_retry(*args, **kwargs):
        raise AssertionError("报告包装器测试不应实际执行代理验证器")

    ftp_probe_with_retry.__report_verifier__ = run_ftp_probe

    recorder = StepRecorder()
    wrapped = attach_cmd_recording_to_closure(
        backend_stub, recorder, ssh_verify
    )
    with recorder.step("FTP真实访问", "验证命令报告入口"):
        wrapped(
            "L5-FTP目录列表",
            ftp_probe_with_retry,
            "ftp_t_wrapper",
            secret,
            2121,
            must_pass=True,
        )

    step = recorder.get_steps()[0]
    assert len(step["verification_commands"]) >= 2
    command = next(
        item["command"] for item in step["verification_commands"]
        if "curl " in item["command"]
    )
    assert '--user "ftp_t_wrapper"' in command
    assert secret not in command
    assert all("验证命令" not in detail for detail in step["details"])
    serialized = json.dumps(step, ensure_ascii=False)
    assert secret not in serialized
    assert "ftp_probe" not in serialized
    assert "if test" not in serialized


def test_l5_manual_commands_match_write_and_connect_fail_operations(backend_stub):
    ftp = _commands(
        backend_stub,
        run_ftp_probe,
        "ftp_t_rw",
        "ftp-secret",
        2121,
        operation="upload_download",
    )
    ftp_text = "\n".join(item["command"] for item in ftp)
    assert "--upload-file" in ftp_text
    assert "--output" in ftp_text
    assert "sha256sum" in ftp_text
    assert "DELE " in ftp_text
    assert any(item["effect"] != "read_only" for item in ftp)

    samba = _commands(
        backend_stub,
        run_samba_probe,
        username="smb_t_rw",
        password="samba-secret",
        host="192.168.148.1",
        iface="ens11",
        operation="upload_download",
        share_name="smb_t_rw_share",
    )
    samba_text = "\n".join(item["command"] for item in samba)
    assert 'ip route get "192.168.148.1"' in samba_text
    assert "smbclient " in samba_text
    assert "put /tmp/" in samba_text
    assert "get ikuai_samba_manual_" in samba_text
    assert "sha256sum" in samba_text
    assert "del ikuai_samba_manual_" in samba_text
    assert any(item["effect"] != "read_only" for item in samba)

    http = _commands(
        backend_stub,
        run_http_probe,
        18080,
        operation="connect_fail",
        host="10.66.0.150",
        iface="enp2s0",
        expected_sha256="a" * 64,
        control_port=18080,
        control_host="192.168.148.1",
        control_iface="ens11",
    )
    http_text = "\n".join(item["command"] for item in http)
    assert 'ip route get "10.66.0.150"' in http_text
    assert 'ip route get "192.168.148.1"' in http_text
    assert "sha256sum" not in http_text
    assert sum("curl " in item["command"] for item in http) == 2


def test_sensitive_command_is_irreversibly_hidden_before_serialization(
    tmp_path,
):
    secret = "never-store-this-command-secret"
    json_secret = "never-store-this-json-secret"
    recorder = StepRecorder()
    with recorder.step("敏感命令防御", "验证报告边界"):
        recorder.add_detail(
            f'【后端数据】{{"community": "{json_secret}"}}'
        )
        recorder.set_actual({"auth_pass": json_secret, "status": "rejected"})
        recorder.add_verification_command({
            "target": "client",
            "target_label": "测试客户端",
            "host": CLIENT_HOST,
            "purpose": "模拟误传敏感命令",
            "command": f"curl --user demo:{secret} http://example.invalid/",
            "expected": "不得展示",
            "contains_secret": True,
            "copy_ready": True,
        })

    data = {
        "total": 1, "passed": 1, "failed": 0, "skipped": 0,
        "total_steps": 1, "duration": "00:00:01",
        "test_cases": [{
            "name": "敏感命令防御", "original_name": "test_sensitive_guard",
            "status": "passed", "duration": "1.00s",
            "steps": recorder.get_steps(), "step_count": 1,
            "error_message": None, "screenshot_path": "",
        }],
    }
    serialized = json.dumps(data, ensure_ascii=False)
    assert secret not in serialized
    assert json_secret not in serialized
    stored = data["test_cases"][0]["steps"][0]["verification_commands"][0]
    assert stored["command"] == "[命令已隐藏：包含敏感信息]"
    assert stored["copy_ready"] is False

    html_path = tmp_path / "sensitive.html"
    ReportGenerator().generate_report(data, str(html_path))
    rendered = html_path.read_text(encoding="utf-8")
    assert secret not in rendered
    assert json_secret not in rendered
    assert "[命令已隐藏：包含敏感信息]" in rendered

    json_path = tmp_path / "sensitive.json"
    excel_path = tmp_path / "sensitive.xlsx"
    json_path.write_text(serialized, encoding="utf-8")
    ok, message = export_results_to_excel(str(json_path), str(excel_path))
    assert ok, message
    from openpyxl import load_workbook
    workbook = load_workbook(excel_path, read_only=True, data_only=False)
    try:
        all_text = "\n".join(
            str(cell.value or "")
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
    finally:
        workbook.close()
    assert secret not in all_text
    assert json_secret not in all_text
    assert "[命令已隐藏：包含敏感信息]" in all_text


def test_step_recorder_marks_pytest_fail_as_failed_instead_of_running():
    recorder = StepRecorder()
    with pytest.raises(pytest.fail.Exception):
        with recorder.step("硬断言失败", "pytest.fail也必须结束步骤"):
            pytest.fail("模拟安全前置失败")

    step = recorder.get_steps()[0]
    assert step["status"] == "failed"
    assert "模拟安全前置失败" in step["error_message"]


def test_step_recorder_preserves_soft_failure_until_step_exit():
    recorder = StepRecorder()
    with recorder.step("操作：提交异常输入；验证：产品明确拒绝"):
        recorder.fail_current_step("产品返回成功但数据库未修改")
        recorder.add_detail("【清理结果】通过：环境已恢复")
    step = recorder.get_steps()[0]
    assert step["status"] == "failed"
    assert "产品返回成功" in step["error_message"]


def test_registered_runtime_secret_is_removed_from_report_text():
    from utils.step_recorder import register_sensitive_value, redact_sensitive_text

    secret = "runtime-only-snmp-credential-X9m!"
    register_sensitive_value(secret)
    text = redact_sensitive_text(f"traceback source: auth_pass={secret}; raw={secret}")
    assert secret not in text
    assert "[已隐藏]" in text


def test_json_html_and_excel_keep_the_exact_same_command(backend_stub, tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    secret = "ftp-report-secret-M7v!"
    source = _commands(
        backend_stub,
        verify_samba_user_database,
        "smb_t_report_ro",
    )[0]
    ftp_source = _commands(
        backend_stub,
        run_ftp_probe,
        "ftp_t_report_ro",
        secret,
        2121,
    )[0]
    recorder = StepRecorder()
    with recorder.step("启用Samba规则", "目标规则: smb_t_report_ro"):
        recorder.add_detail("【页面验证】✓ 规则状态已变为启用")
        recorder.add_detail("【后端验证·L1】✓ 数据库字段正确")
        recorder.add_detail("【后端数据】{\"enabled\": \"yes\"}")
        recorder.add_verification_commands([source, ftp_source])

    data = {
        "total": 1,
        "passed": 1,
        "failed": 0,
        "skipped": 0,
        "total_steps": 1,
        "duration": "00:00:01",
        "test_cases": [
            {
                "name": "Samba服务完整测试",
                "original_name": "test_samba_server_comprehensive",
                "status": "passed",
                "duration": "1.00s",
                "steps": recorder.get_steps(),
                "step_count": 1,
                "error_message": None,
                "screenshot_path": "",
            }
        ],
    }
    json_path = tmp_path / "test_results.json"
    json_text = json.dumps(data, ensure_ascii=False, indent=2)
    json_path.write_text(json_text, encoding="utf-8")
    loaded = json.loads(json_path.read_text(encoding="utf-8"))

    html_path = tmp_path / "report.html"
    ReportGenerator().generate_report(
        loaded,
        str(html_path),
        report_title="本地服务自动化测试报告",
        device_info={"ip": ROUTER_HOST, "username": "admin"},
    )
    rendered_html = html_path.read_text(encoding="utf-8")
    html_commands = [
        html_lib.unescape(item)
        for item in re.findall(
            r'<code id="verification-command-[^"]+">(.*?)</code>',
            rendered_html,
            flags=re.DOTALL,
        )
    ]

    excel_path = tmp_path / "report.xlsx"
    success, message = export_results_to_excel(str(json_path), str(excel_path))
    assert success, message
    workbook = load_workbook(excel_path, read_only=True, data_only=False)
    try:
        command_sheet = workbook["复验命令"]
        headers = [cell.value for cell in command_sheet[1]]
        command_column = headers.index("命令") + 1
        excel_commands = [
            command_sheet.cell(row=row, column=command_column).value
            for row in range(2, command_sheet.max_row + 1)
            if command_sheet.cell(row=row, column=command_column).value is not None
        ]
        workbook_text = "\n".join(
            str(value)
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
    finally:
        workbook.close()

    json_commands = [
        item["command"]
        for item in loaded["test_cases"][0]["steps"][0][
            "verification_commands"
        ]
    ]
    assert json_commands == [source["command"], ftp_source["command"]]
    assert html_commands == json_commands
    assert excel_commands == json_commands
    assert secret not in json_text
    assert secret not in rendered_html
    assert secret not in workbook_text

    # 中文卡片和复制按钮可见，复制逻辑只读取 code.textContent。
    assert "人工复验命令" in rendered_html
    assert "用途:" in rendered_html
    assert "预期:" in rendered_html
    assert "复制命令" in rendered_html
    assert "var commandText = commandElement.textContent;" in rendered_html
    assert "【页面验证】" in rendered_html
    assert "【后端验证·L1】" in rendered_html


def test_snmp_router_commands_cover_l1_l4_without_exposing_secrets(backend_stub):
    secrets = {
        "snmp-community-W7q!",
        "snmp-auth-K4r!",
        "snmp-priv-P9m!",
    }
    result = SimpleNamespace(
        passed=True,
        message="SNMP运行时一致，community=must-not-survive",
        details={"listen_port": 1161},
    )
    commands = build_verification_commands(
        backend_stub,
        verify_snmp_runtime_consistency,
        kwargs={
            "expected_fields": {
                "enabled": "yes",
                "listen_port": 1161,
                "version": "v3",
            },
            "expected_secrets": {
                "community": "snmp-community-W7q!",
                "auth_pass": "snmp-auth-K4r!",
                "priv_pass": "snmp-priv-P9m!",
            },
        },
        result=result,
    )

    assert commands
    assert all(item["target"] == "router" for item in commands)
    command_text = "\n".join(item["command"] for item in commands)
    payload = json.dumps(commands, ensure_ascii=False)
    for secret in secrets | {"must-not-survive"}:
        assert secret not in payload
    assert "FROM snmp_conf" in command_text
    assert "community_state" in command_text
    assert "auth_pass_state" in command_text
    assert "priv_pass_state" in command_text
    assert "/usr/ikuai/script/netsnmp.sh" in command_text
    assert "/var/run/snmp/snmpd.conf" in command_text
    assert "cat /var/run/snmp/snmpd.conf" not in command_text
    assert "rocommunity [hidden]" in command_text
    assert "pidof snmpd" in command_text
    assert "pidof ik_snmp_subagent" in command_text
    assert ":1161[[:space:]]" in command_text
    assert "grep -i ':0489 ' /proc/net/udp6" in command_text
    assert "deny 1161 0.0.0.0/0 0-65535" in command_text
    assert "FROM upnpd_conf" in command_text
    assert any("UPnP关闭时无输出且该项不适用" in item["expected"] for item in commands)
    assert "iptables-save" in command_text
    assert "ipset list -n" in command_text
    assert all(item["actual"].startswith("通过：") for item in commands)
    for item in commands:
        _assert_copy_ready(item)


def test_snmp_singleton_contract_and_client_routes_are_copy_ready(backend_stub):
    contract = _commands(backend_stub, verify_snmp_singleton_contract)
    assert all(item["target"] == "router" for item in contract)
    contract_text = "\n".join(item["command"] for item in contract)
    assert "sqlite3 /etc/mnt/ikuai/config.db '.schema snmp_conf'" in contract_text
    assert "SELECT count(*) AS row_count FROM snmp_conf" in contract_text
    assert (
        "grep -nF 'url=advanced-service/snmpd-config' "
        "/usr/ikuai/script/netsnmp.sh"
    ) in contract_text

    wan_route = _commands(
        backend_stub,
        verify_snmp_client_route,
        "10.66.0.150:1161",
        "enp2s0",
    )
    assert len(wan_route) == 1
    assert wan_route[0]["target"] == "client"
    assert wan_route[0]["command"] == "ip route get 10.66.0.150"
    assert "dev enp2s0" in wan_route[0]["expected"]

    lan_probe = _commands(
        backend_stub,
        run_snmp_probe,
        version="v2c",
        host="192.168.148.1:2161",
        oid="1.3.6.1.2.1.1.1.0",
        community="route-secret-never-report",
    )
    lan_route = next(item for item in lan_probe if item["command"].startswith("ip route get "))
    helper = next(
        item for item in lan_probe
        if item["command"].startswith(
            "sudo -n /usr/local/sbin/ikuai-snmp-verify "
        )
    )
    assert lan_route["command"] == "ip route get 192.168.148.1"
    assert "dev ens11" in lan_route["expected"]
    assert '--host "192.168.148.1:2161"' in helper["command"]
    assert "route-secret-never-report" not in json.dumps(
        lan_probe, ensure_ascii=False
    )


@pytest.mark.parametrize(
    ("kwargs", "mode", "extra"),
    [
        (
            {
                "version": "v2c",
                "host": ROUTER_HOST,
                "oid": "1.3.6.1.2.1.1.1.0",
                "operation": "get",
                "community": "v2c-secret-Q2x!",
            },
            "v2c",
            "",
        ),
        (
            {
                "version": "v3",
                "host": ROUTER_HOST,
                "oid": "1.3.6.1.2.1.1",
                "operation": "walk",
                "username": "private-user",
                "security": "authPriv",
                "auth_proto": "SHA",
                "auth_pass": "v3-auth-secret-L8z!",
                "priv_proto": "AES",
                "priv_pass": "v3-priv-secret-B6n!",
            },
            "v3-priv",
            "--auth-proto SHA --priv-proto AES",
        ),
    ],
)
def test_snmp_l5_uses_real_permission_controlled_helper(
    backend_stub, kwargs, mode, extra
):
    commands = _commands(backend_stub, run_snmp_probe, **kwargs)
    helper = next(
        item for item in commands
        if item["command"].startswith(
            "sudo -n /usr/local/sbin/ikuai-snmp-verify "
        )
    )
    command = helper["command"]
    assert helper["target"] == "client"
    assert helper["host"] == CLIENT_HOST
    assert helper["interactive"] is True
    assert helper["effect"] != "read_only"
    assert f"--mode {mode}" in command
    assert f"--operation {kwargs['operation']}" in command
    assert f'--host "{ROUTER_HOST}"' in command
    assert f'--oid "{kwargs["oid"]}"' in command
    if extra:
        assert extra in command
    serialized = json.dumps(commands, ensure_ascii=False)
    for secret_key in ("community", "username", "auth_pass", "priv_pass"):
        secret = kwargs.get(secret_key)
        if secret:
            assert secret not in serialized
    assert "snmpget" not in command and "snmpwalk" not in command
    assert SHELL_SCRIPT_TOKENS.search(command) is None
    assert any(
        item["command"] == (
            "stat -c '%U:%G %a %n' /usr/local/sbin/ikuai-snmp-verify"
        )
        for item in commands
    )
    assert any(
        item["command"] == (
            "find /tmp -maxdepth 1 -name 'ikuai-snmp-verify.*' -print"
        )
        for item in commands
    )


def test_snmp_wrapper_hides_internal_stdin_script_and_keeps_actual(backend_stub):
    secret = "wrapper-snmp-community-H7m!"
    backend_stub.mark_cmd_start = lambda: (0, 0)
    backend_stub.collect_cmds_since_mark = lambda mark: [
        f"[client] community={secret}; rc=$?; echo __SNMP_INTERNAL__"
    ]

    def ssh_verify(label, verify_func, *args, must_pass=False, **kwargs):
        return SimpleNamespace(
            passed=True,
            message="V2C get返回目标OID和值",
            details={},
            raw_output="",
        )

    def snmp_probe_with_retry(*args, **kwargs):
        raise AssertionError("报告包装器测试不应执行代理验证器")

    snmp_probe_with_retry.__report_verifier__ = run_snmp_probe
    recorder = StepRecorder()
    wrapped = attach_cmd_recording_to_closure(
        backend_stub, recorder, ssh_verify
    )
    with recorder.step("SNMP真实协议", "确认安全报告入口"):
        wrapped(
            "L5-SNMP-v2c-get",
            snmp_probe_with_retry,
            "v2c",
            ROUTER_HOST,
            "1.3.6.1.2.1.1.1.0",
            community=secret,
            must_pass=True,
        )

    step = recorder.get_steps()[0]
    payload = json.dumps(step, ensure_ascii=False)
    assert secret not in payload
    assert "__SNMP_INTERNAL__" not in payload
    assert "rc=$?" not in payload
    helper = next(
        item for item in step["verification_commands"]
        if item["command"].startswith(
            "sudo -n /usr/local/sbin/ikuai-snmp-verify "
        )
    )
    assert helper["target"] == "client"
    assert helper["actual"] == "通过：V2C get返回目标OID和值"


def test_snmp_negative_manual_command_keeps_failure_reason_outside_command(
    backend_stub,
):
    commands = _commands(
        backend_stub,
        run_snmp_probe,
        "v3",
        "192.168.148.1:2161",
        "1.3.6.1.2.1.1.5.0",
        username="safe-user-label",
        security="authPriv",
        auth_proto="SHA",
        auth_pass="never-report-auth",
        priv_proto="AES",
        priv_pass="never-report-priv",
        expect_success=False,
        expected_failure="privacy",
    )
    protocol = next(item for item in commands if "真实SNMP" in item["purpose"])
    assert protocol["target"] == "client"
    assert "expected_failure" not in protocol["command"]
    assert "never-report-auth" not in json.dumps(commands, ensure_ascii=False)
    assert "never-report-priv" not in json.dumps(commands, ensure_ascii=False)
    assert protocol["copy_ready"] is True
    assert protocol["contains_secret"] is False
    assert protocol["interactive"] is True
    assert SHELL_SCRIPT_TOKENS.search(protocol["command"]) is None


def test_snmp_actual_and_chinese_sections_match_json_html_excel(
    backend_stub, tmp_path
):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    commands = build_verification_commands(
        backend_stub,
        run_snmp_probe,
        kwargs={
            "version": "v2c",
            "host": ROUTER_HOST,
            "oid": "1.3.6.1.2.1.1.1.0",
            "community": "never-report-community",
        },
        result=SimpleNamespace(passed=True, message="OID和值匹配"),
    )
    protocol = next(
        item for item in commands
        if item["command"].startswith(
            "sudo -n /usr/local/sbin/ikuai-snmp-verify "
        )
    )
    recorder = StepRecorder()
    with recorder.step("操作：执行SNMP协议；验证：页面、运行时和协议一致"):
        recorder.add_detail("【测试操作】执行v2c snmpget")
        recorder.add_detail("【页面验证】通过：保存反馈正确")
        recorder.add_detail("【后端验证·L1】通过：snmp_conf字段正确")
        recorder.add_detail("【运行时验证】通过：UDP监听正确")
        recorder.add_detail("【协议验证】通过：OID和值匹配")
        recorder.add_detail("【清理结果】通过：临时文件无残留")
        recorder.add_detail("【不适用】无专用iptables/ipset")
        recorder.set_actual("协议返回目标OID和值")
        recorder.add_verification_command(protocol)

    data = {
        "total": 1,
        "passed": 1,
        "failed": 0,
        "skipped": 0,
        "total_steps": 1,
        "duration": "00:00:01",
        "test_cases": [{
            "name": "SNMP服务完整测试",
            "original_name": "test_snmp_server_comprehensive",
            "status": "passed",
            "duration": "1.00s",
            "steps": recorder.get_steps(),
            "step_count": 1,
            "error_message": None,
            "screenshot_path": "",
        }],
    }
    json_path = tmp_path / "snmp.json"
    json_text = json.dumps(data, ensure_ascii=False, indent=2)
    json_path.write_text(json_text, encoding="utf-8")
    html_path = tmp_path / "snmp.html"
    ReportGenerator().generate_report(data, str(html_path))
    html_text = html_path.read_text(encoding="utf-8")
    excel_path = tmp_path / "snmp.xlsx"
    ok, message = export_results_to_excel(str(json_path), str(excel_path))
    assert ok, message

    workbook = load_workbook(excel_path, read_only=True, data_only=False)
    try:
        command_sheet = workbook["复验命令"]
        command_headers = [cell.value for cell in command_sheet[1]]
        actual_column = command_headers.index("实际") + 1
        excel_actual = command_sheet.cell(2, actual_column).value
        workbook_text = "\n".join(
            str(value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
    finally:
        workbook.close()

    assert protocol["actual"] == "通过：OID和值匹配"
    assert protocol["actual"] in json_text
    assert protocol["actual"] in html_text
    assert excel_actual == protocol["actual"]
    assert "协议返回目标OID和值" in json_text
    assert "协议返回目标OID和值" in html_text
    assert "协议返回目标OID和值" in workbook_text
    for css_class in (
        "detail-test-action",
        "detail-page-verification",
        "detail-backend-verification",
        "detail-runtime-verification",
        "detail-protocol-verification",
        "detail-cleanup-result",
        "detail-not-applicable",
    ):
        assert css_class in html_text
    assert "never-report-community" not in json_text
    assert "never-report-community" not in html_text
    assert "never-report-community" not in workbook_text
