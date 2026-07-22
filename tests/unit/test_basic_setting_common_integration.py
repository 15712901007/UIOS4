"""基础设置公共接入、复验命令和归档审计的离线回归。"""

from __future__ import annotations

import copy
import importlib.util
import json
import os
import re
from pathlib import Path
from types import SimpleNamespace

import pytest

from utils.backend_verifier import BackendVerifier
from utils.basic_setting_artifact_audit import audit_basic_setting_artifacts
from utils.replay_commands import build_verification_commands
from utils.report_generator import ReportGenerator
from utils.step_recorder import StepRecorder
from utils.test_results_to_excel import export_results_to_excel
from utils.verify_helper import attach_cmd_recording_to_closure


ROOT = Path(__file__).resolve().parents[2]
PRIVATE_HOSTNAME = "basic-private-host-Q7x"
PRIVATE_NTP = "private-ntp-Q7x.invalid"
PRIVATE_PASSWORDS = (
    "device-password-Q7x!",
    "router-password-Q7x!",
    "console-password-Q7x!",
    "client-password-Q7x!",
)
PRIVATE_USERNAMES = (
    "device-user-Q7x",
    "router-user-Q7x",
    "console-user-Q7x",
    "client-user-Q7x",
)
SHELL_SCRIPT_TOKENS = re.compile(
    r"\b(?:if|then|fi|for|do|done|while|case|esac)\b|"
    r"\$\(|\$\{|\$(?:\?|[A-Za-z_][A-Za-z0-9_]*)|"
    r"(?:^|[ ;])rc\s*=|__[A-Z0-9_]+__|\bbase64(?:\s|$)|"
    r"(?:^|[ ;])\[(?:router|client)\](?:$|[ ;])"
)


@pytest.fixture(scope="module")
def project_conftest():
    spec = importlib.util.spec_from_file_location(
        "_basic_setting_project_conftest", ROOT / "tests" / "conftest.py"
    )
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


@pytest.fixture
def backend_stub():
    backend = object.__new__(BackendVerifier)
    backend._ssh_config = SimpleNamespace(
        router=SimpleNamespace(host="192.0.2.1"),
        client=SimpleNamespace(host="192.0.2.2"),
        iperf3_server="198.51.100.40",
    )
    return backend


def _snapshot():
    return {
        "row": {
            "id": 1,
            "hostname": PRIVATE_HOSTNAME,
            "ntpserver_list": PRIVATE_NTP,
            "switch_nat": 1,
            "link_mode": 0,
            "fast_nat": 1,
        },
        "client": {
            "server_ip": "198.51.100.40",
            "route": "198.51.100.40 via 192.168.148.1 dev ens11",
        },
    }


def _verifier_calls(backend):
    snapshot = _snapshot()
    return [
        (backend.verify_basic_singleton_contract, (), {}),
        (
            backend.verify_basic_database,
            (),
            {
                "expected_fields": {
                    "hostname": PRIVATE_HOSTNAME,
                    "ntpserver_list": PRIVATE_NTP,
                    "switch_nat": 1,
                },
                "snapshot": snapshot,
            },
        ),
        (
            backend.verify_basic_generated_state,
            (),
            {"expected_fields": {"hostname": PRIVATE_HOSTNAME}},
        ),
        (backend.verify_basic_nat_runtime, (1, 0), {}),
        (backend.verify_basic_link_runtime, (0,), {}),
        (backend.verify_basic_link_topology_safety, (), {}),
        (backend.verify_basic_acceleration_runtime, (1,), {}),
        (backend.verify_basic_ntp_runtime, (1, 0, 60), {}),
        (
            backend.verify_basic_runtime_consistency,
            ({"switch_nat": 1},),
            {"snapshot": snapshot},
        ),
        (
            backend.verify_basic_reinit,
            ({"switch_nat": 1},),
            {"snapshot": snapshot},
        ),
        (backend.verify_basic_environment_unchanged, (snapshot,), {}),
        (backend.verify_basic_management_health, (), {}),
        (backend.verify_basic_test_artifacts_absent, (snapshot,), {}),
        (backend.verify_basic_client_route, (), {}),
        (backend.run_basic_iperf_probe, (2, True), {}),
        (backend.run_basic_acceleration_probe, (2,), {}),
        (backend.run_basic_route_mode_probe, (), {}),
        (backend.run_basic_fullcone_probe, (True,), {}),
        (backend.run_basic_ntp_protocol_probe, (), {}),
    ]


def _all_basic_commands(backend):
    result = SimpleNamespace(
        passed=True,
        message="基础设置验证通过，私有字段只核对长度和布尔状态",
        details={},
    )
    commands = []
    for verifier, args, kwargs in _verifier_calls(backend):
        built = build_verification_commands(
            backend, verifier, args=args, kwargs=kwargs, result=result
        )
        assert built is not None
        commands.extend(built)
    return commands


def test_basic_ntp_time_flow_matches_real_conditional_contract():
    test_source = (
        ROOT / "tests" / "device_setting" / "test_basic_setting_comprehensive.py"
    ).read_text(encoding="utf-8")
    page_source = (
        ROOT / "pages" / "device_setting" / "basic_setting_page.py"
    ).read_text(encoding="utf-8")

    # 非法日期必须先证明组件保留了输入，不能在组件已归一化时误点保存。
    assert "manual_value_retained" in test_source
    assert 'page.field_matches("manual_time", invalid_manual_time)' in test_source
    assert "manual_prepared and manual_value_retained and not manual_errors" in test_source
    assert 'observation.get("present") and observation.get("visible")' in test_source
    assert "不能宣称产品拒绝" in test_source
    assert "同步周期原生边界5至240" in test_source
    assert "allow_native_input_rejection" in test_source
    assert 'observation.get("control") == "number"' in test_source
    assert 'before_observation.get("length")' in test_source
    assert '("同步周期非法字符", "abc", True)' in test_source

    # set_time 的真实 UI 合同会进入手动模式；立即对时前先恢复条件字段，
    # 数据库断言仅允许 NTP 字段变化而不是错误要求整行不变。
    assert 'explicit={"switch_ntp": 0, "ntpserver_list": ""}' in test_source
    assert 'page.select_ntp_config("builtin")' in test_source
    assert test_source.index("custom_set = page.set_manual_time") < test_source.index(
        "saved_custom = save_form"
    ) < test_source.index("custom_sync = page.sync_time_now")

    # 条件渲染会留下隐藏旧控件，Page Object 必须优先返回可见原生节点。
    field_block = page_source.split("    def _field(self, name: str)", 1)[1].split(
        "    def _field_item", 1
    )[0]
    assert "node.is_visible()" in field_block
    assert "first_native" in field_block
    assert "first_visible" in field_block
    assert "elif node.is_visible()" not in field_block


def test_basic_commands_are_copy_ready_complete_and_private(backend_stub):
    commands = _all_basic_commands(backend_stub)
    assert commands
    serialized = json.dumps(commands, ensure_ascii=False)
    for private_value in (PRIVATE_HOSTNAME, PRIVATE_NTP, *PRIVATE_PASSWORDS):
        assert private_value not in serialized
    assert not re.search(r"(?<![A-Za-z0-9])[A-Za-z]:[\\/]", serialized)
    assert "restore basic-setting" not in serialized

    for item in commands:
        assert item["target"] in {"router", "client"}
        assert item["copy_ready"] is True
        assert item["contains_secret"] is False
        assert item["command"] == item["command"].strip()
        assert "\n" not in item["command"] and "\r" not in item["command"]
        assert "[router]" not in item["command"]
        assert "[client]" not in item["command"]
        assert SHELL_SCRIPT_TOKENS.search(item["command"]) is None
        for key in ("purpose", "expected", "actual", "valid_when", "effect"):
            assert str(item.get(key, "")).strip(), (key, item["purpose"])

    command_text = "\n".join(item["command"] for item in commands)
    assert "pgrep -af 'python3 /tmp/[i]kuai-basic-'" in command_text
    assert "ip -4 route show table main" in command_text
    assert "ip -4 rule show" in command_text
    assert "hwclock -r" in command_text
    assert "sha256sum /etc/passwd" in command_text

    route_mode = build_verification_commands(
        backend_stub,
        backend_stub.run_basic_route_mode_probe,
        kwargs={
            "server_ip": "198.51.100.40",
            "lan_ip": "192.168.148.2",
            "wan_iface": "wan1",
            "client_iface": "ens11",
        },
        result=SimpleNamespace(
            passed=True, message="WAN侧源地址保持不变", details={}
        ),
    )
    route_text = "\n".join(item["command"] for item in route_mode)
    assert "timeout -t 8 tcpdump -ni wan1 -c 4" in route_text
    assert "ping -I ens11 -c 4 -i 0.25 -W 2 198.51.100.40" in route_text
    assert "iperf3" not in route_text
    ping_command = next(item for item in route_mode if item["command"].startswith("ping "))
    assert "允许无回程" in ping_command["expected"]


def test_basic_restore_prepare_cleanup_never_expose_internal_commands(
    backend_stub,
):
    snapshot = _snapshot()

    def cleanup_basic_setting(snapshot=None):
        raise AssertionError("离线命令生成不应执行清理函数")

    for verifier, args in (
        (backend_stub.get_basic_environment_snapshot, ()),
        (backend_stub.prepare_basic_l5_route, ()),
        (backend_stub.restore_basic_environment, (snapshot,)),
        (cleanup_basic_setting, (snapshot,)),
    ):
        assert build_verification_commands(
            backend_stub, verifier, args=args
        ) == []

    internal_secret = "internal-restore-secret-Q7x!"
    backend_stub.mark_cmd_start = lambda: (0, 0)
    backend_stub.collect_cmds_since_mark = lambda mark: [
        "[router] set -eu; rc=$?; " + internal_secret
    ]

    def ssh_verify(label, verify_func, *args, must_pass=False, **kwargs):
        return SimpleNamespace(passed=True, message="恢复完成", details={})

    recorder = StepRecorder()
    wrapped = attach_cmd_recording_to_closure(
        backend_stub, recorder, ssh_verify
    )
    with recorder.step("操作：执行finally恢复；验证：内部脚本不进入报告"):
        wrapped(
            "基础设置恢复",
            backend_stub.restore_basic_environment,
            snapshot,
            must_pass=True,
        )
    payload = json.dumps(recorder.get_steps(), ensure_ascii=False)
    assert internal_secret not in payload
    assert "set -eu" not in payload
    assert "rc=$?" not in payload
    assert recorder.get_steps()[0]["verification_commands"] == []

    # DNS加速的历史 ``verify_dns_basic_full_chain`` 不是设备基础设置，
    # 必须继续走原有命令路径，不能被名称中的basic误判为已迁移。
    assert build_verification_commands(
        backend_stub, backend_stub.verify_dns_basic_full_chain
    ) is None


def test_basic_verify_helper_records_manual_commands_not_raw_scripts(
    backend_stub,
):
    raw_secret = "raw-basic-script-secret-Q7x!"
    backend_stub.mark_cmd_start = lambda: (0, 0)
    backend_stub.collect_cmds_since_mark = lambda mark: [
        "[router] __BASIC_INTERNAL__ password=" + raw_secret
    ]

    def ssh_verify(label, verify_func, *args, must_pass=False, **kwargs):
        return SimpleNamespace(
            passed=True, message="基础设置数据库字段一致", details={}
        )

    recorder = StepRecorder()
    wrapped = attach_cmd_recording_to_closure(
        backend_stub, recorder, ssh_verify
    )
    with recorder.step("操作：执行L1验证；验证：只记录人工命令"):
        wrapped(
            "L1",
            backend_stub.verify_basic_database,
            expected_fields={"switch_nat": 1},
            must_pass=True,
        )
    step = recorder.get_steps()[0]
    payload = json.dumps(step, ensure_ascii=False)
    assert step["verification_commands"]
    assert raw_secret not in payload
    assert "__BASIC_INTERNAL__" not in payload


def test_basic_common_fixture_mapping_markers_and_secret_registration(
    project_conftest, monkeypatch
):
    fake_config = SimpleNamespace(
        device=SimpleNamespace(
            username=PRIVATE_USERNAMES[0], password=PRIVATE_PASSWORDS[0]
        ),
        ssh=SimpleNamespace(
            router=SimpleNamespace(
                username=PRIVATE_USERNAMES[1],
                password=PRIVATE_PASSWORDS[1],
                console_username=PRIVATE_USERNAMES[2],
                console_password=PRIVATE_PASSWORDS[2],
            ),
            client=SimpleNamespace(
                username=PRIVATE_USERNAMES[3], password=PRIVATE_PASSWORDS[3]
            ),
        ),
    )
    registered = []
    monkeypatch.setattr(
        project_conftest, "get_config_with_env", lambda: fake_config
    )
    monkeypatch.setattr(
        project_conftest,
        "register_sensitive_values",
        lambda values: registered.extend(values),
    )

    assert project_conftest.config.__wrapped__() is fake_config
    assert tuple(registered) == (
        PRIVATE_USERNAMES[0], PRIVATE_PASSWORDS[0],
        PRIVATE_USERNAMES[1], PRIVATE_PASSWORDS[1],
        PRIVATE_USERNAMES[2], PRIVATE_PASSWORDS[2],
        PRIVATE_USERNAMES[3], PRIVATE_PASSWORDS[3],
    )
    assert (
        project_conftest.TEST_NAME_MAPPING["test_basic_setting_comprehensive"]
        == "设备设置-基础设置"
    )
    assert hasattr(project_conftest, "basic_setting_page")
    assert hasattr(project_conftest, "basic_setting_page_logged_in")
    ini_text = (ROOT / "pytest.ini").read_text(encoding="utf-8")
    assert "device_setting: 设备设置模块测试" in ini_text
    assert "basic_setting: 设备设置-基础设置模块测试" in ini_text


def test_basic_session_report_hides_username_and_logs_no_absolute_paths(
    project_conftest, monkeypatch, tmp_path, capsys
):
    output_dir = tmp_path / "private-output" / "reports"
    screenshot_dir = tmp_path / "private-output" / "screenshots"
    private_username = "private-device-user-Q7x"
    fake_config = SimpleNamespace(
        device=SimpleNamespace(ip="192.0.2.1", username=private_username),
        report=SimpleNamespace(
            output_dir=str(output_dir),
            screenshot_dir=str(screenshot_dir),
            tester="自动化测试",
            version="v4.0",
        ),
    )
    original_results = copy.deepcopy(project_conftest._test_results)
    project_conftest._test_results.clear()
    project_conftest._test_results.update({
        "total": 1,
        "passed": 1,
        "failed": 0,
        "skipped": 0,
        "total_steps": 1,
        "duration": "00:00:01",
        "start_time": None,
        "end_time": None,
        "test_cases": [{
            "name": "设备设置-基础设置",
            "original_name": "test_basic_setting_comprehensive",
            "status": "passed",
            "duration": "0.01s",
            "steps": [],
            "step_count": 0,
            "error_message": None,
            "screenshot": None,
        }],
    })
    monkeypatch.setattr(project_conftest, "get_config", lambda: fake_config)
    monkeypatch.setattr(
        project_conftest, "get_registered_sensitive_values", lambda: []
    )
    monkeypatch.setattr(
        project_conftest, "clear_registered_sensitive_values", lambda: None
    )
    monkeypatch.delenv("TESTER", raising=False)
    monkeypatch.delenv("TEST_VERSION", raising=False)
    session = SimpleNamespace(exitstatus=0)
    try:
        project_conftest.pytest_sessionfinish(session, 0)
    finally:
        project_conftest._test_results.clear()
        project_conftest._test_results.update(original_results)

    html_path = next(output_dir.glob("test_report_*.html"))
    html_text = html_path.read_text(encoding="utf-8")
    stdout = capsys.readouterr().out
    assert private_username not in html_text
    assert private_username not in stdout
    assert str(tmp_path) not in stdout
    assert "SNMP协议秘密" not in stdout
    assert "内存登记敏感值" in stdout


def test_basic_setting_failure_never_embeds_page_screenshot(project_conftest):
    screenshot_calls = []

    class PageStub:
        def is_closed(self):
            return False

        def screenshot(self, **kwargs):
            screenshot_calls.append(kwargs)

    page_holder = SimpleNamespace(page=PageStub())
    item = SimpleNamespace(
        funcargs={"basic_setting_page_logged_in": page_holder},
        name="test_basic_setting_comprehensive",
        get_closest_marker=lambda name: object() if name == "basic_setting" else None,
    )
    call = SimpleNamespace(when="call")
    report = SimpleNamespace(failed=True)
    outcome = SimpleNamespace(get_result=lambda: report)

    hook = project_conftest.pytest_runtest_makereport(item, call)
    next(hook)
    with pytest.raises(StopIteration):
        hook.send(outcome)

    assert screenshot_calls == []
    assert not hasattr(report, "extra")


def test_basic_screenshot_json_path_is_relative(project_conftest, tmp_path):
    project_root = tmp_path / "project"
    screenshot_dir = project_root / "reports" / "screenshots"
    output_dir = project_root / "reports" / "output"
    screenshot_dir.mkdir(parents=True)
    output_dir.mkdir(parents=True)
    screenshot_name = (
        "test_basic_setting_comprehensive_20260716_120000_failure.png"
    )
    (screenshot_dir / screenshot_name).write_bytes(b"png")
    results = {
        "total": 1,
        "passed": 0,
        "failed": 1,
        "skipped": 0,
        "test_cases": [{
            "name": "设备设置-基础设置",
            "original_name": "test_basic_setting_comprehensive",
            "status": "failed",
            "duration": "0.01s",
            "steps": [],
            "step_count": 0,
            "error_message": "失败",
            "screenshot": "data:image/png;base64,hidden",
        }],
    }
    json_path = project_conftest._dump_test_results_json(
        results, str(output_dir), str(screenshot_dir), str(project_root)
    )
    data = json.loads(Path(json_path).read_text(encoding="utf-8"))
    screenshot_path = data["test_cases"][0]["screenshot_path"]
    assert screenshot_path == "reports/screenshots/" + screenshot_name
    assert not os.path.isabs(screenshot_path)
    assert not re.match(r"^[A-Za-z]:[\\/]", screenshot_path)

    external_dir = tmp_path / "private-external-screenshots"
    external_dir.mkdir()
    (external_dir / screenshot_name).write_bytes(b"png")
    external_label = project_conftest._find_screenshot_path(
        str(external_dir),
        "test_basic_setting_comprehensive",
        str(project_root),
    )
    assert external_label == screenshot_name
    assert "private-external" not in external_label


def _artifact_data(commands):
    statuses = ("passed", "failed", "warning", "skipped")
    steps = []
    for index, status in enumerate(statuses, 1):
        steps.append({
            "name": f"操作：生成四态步骤{index}；验证：三类产物一致",
            "description": "基础设置归档审计夹具",
            "status": status,
            "duration": "0.01s",
            "details": [
                f"{section}\n不适用：离线审计夹具"
                for section in (
                    "【测试操作】", "【页面验证】", "【后端验证】",
                    "【运行时验证】", "【协议验证】", "【清理结果】",
                )
            ],
            "verification_commands": commands if index == 1 else [],
            "actual": f"步骤{index}已记录",
            "error_message": "离线失败态证据" if status == "failed" else None,
        })
    return {
        "schema_version": 2,
        "total": 1,
        "passed": 0,
        "failed": 1,
        "skipped": 0,
        "total_steps": len(steps),
        "duration": "00:00:01",
        "test_cases": [{
            "name": "设备设置-基础设置",
            "original_name": "test_basic_setting_comprehensive",
            "status": "failed",
            "duration": "1.00s",
            "steps": steps,
            "step_count": len(steps),
            "error_message": "离线失败态证据",
            "error_traceback": "完整源码堆栈已按凭据安全策略隐藏。",
            "screenshot_path": "reports/screenshots/basic_failure.png",
        }],
    }


def _write_artifacts(tmp_path, data):
    json_path = tmp_path / "basic.json"
    html_path = tmp_path / "basic.html"
    excel_path = tmp_path / "basic.xlsx"
    json_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    ReportGenerator().generate_report(data, str(html_path))
    success, message = export_results_to_excel(str(json_path), str(excel_path))
    assert success, message
    return json_path, html_path, excel_path


def _basic_report_html_for_smoke(tmp_path, backend):
    configured = (
        os.environ.get("BASIC_SETTING_REAL_HTML")
        or os.environ.get("BASIC_SETTING_REPORT_HTML")
        or os.environ.get("IKUAI_BASIC_SETTING_REPORT_HTML")
    )
    if configured:
        report_path = Path(configured).expanduser()
        assert report_path.is_file(), "环境变量指定的基础设置HTML不存在"
        return report_path.resolve()
    commands = _all_basic_commands(backend)[:3]
    return _write_artifacts(tmp_path, _artifact_data(commands))[1].resolve()


def test_basic_report_smoke_accepts_existing_real_html_environment(
    backend_stub, tmp_path, monkeypatch
):
    real_report = tmp_path / "archived-basic-report.html"
    real_report.write_text("<html><body>基础设置真实报告占位</body></html>", encoding="utf-8")
    monkeypatch.setenv("BASIC_SETTING_REAL_HTML", str(real_report))
    assert _basic_report_html_for_smoke(tmp_path, backend_stub) == real_report.resolve()


def test_basic_report_desktop_mobile_and_every_copy_button(
    backend_stub, tmp_path
):
    pytest.importorskip("playwright")
    from playwright.sync_api import sync_playwright

    report_path = _basic_report_html_for_smoke(tmp_path, backend_stub)
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()
        page.add_init_script("""(() => {
            Object.defineProperty(window, 'isSecureContext', {
                configurable: true,
                value: true
            });
            Object.defineProperty(navigator, 'clipboard', {
                configurable: true,
                value: {
                    writeText: text => {
                        window.__basicCopiedText = text;
                        return Promise.resolve();
                    }
                }
            });
        })()""")

        for width, height in ((1440, 900), (390, 844)):
            page.set_viewport_size({"width": width, "height": height})
            page.goto(report_path.as_uri(), wait_until="load")
            page.evaluate("""() => {
                document.querySelectorAll('.test-case-details').forEach(
                    element => element.classList.add('show'));
                document.querySelectorAll('details').forEach(
                    element => element.open = true);
            }""")
            layout = page.evaluate("""() => ({
                pageFits: document.documentElement.scrollWidth <= window.innerWidth + 1,
                cardsFit: [...document.querySelectorAll(
                    '.container,.test-case,.test-case-details,' +
                    '.verification-command-card,.copy-verification-command'
                )].every(element => {
                    const rect = element.getBoundingClientRect();
                    return rect.left >= -1 && rect.right <= window.innerWidth + 1;
                })
            })""")
            assert layout == {"pageFits": True, "cardsFit": True}

        page.set_viewport_size({"width": 1440, "height": 900})
        page.goto(report_path.as_uri(), wait_until="load")
        page.evaluate("""() => {
            document.querySelectorAll('.test-case-details').forEach(
                element => element.classList.add('show'));
            document.querySelectorAll('details').forEach(
                element => element.open = true);
        }""")
        buttons = page.locator("button.copy-verification-command")
        assert buttons.count() > 0
        for index in range(buttons.count()):
            button = buttons.nth(index)
            code = button.locator(
                "xpath=ancestor::div[contains(@class,'verification-command-card')][1]//code"
            ).text_content()
            assert "[router]" not in code and "[client]" not in code
            button.click()
            page.wait_for_function(
                "text => window.__basicCopiedText === text", arg=code
            )
            assert page.evaluate("window.__basicCopiedText") == code

        context.close()
        browser.close()


def test_basic_artifact_audit_checks_unique_case_four_states_and_consistency(
    backend_stub, tmp_path
):
    commands = _all_basic_commands(backend_stub)
    paths = _write_artifacts(tmp_path, _artifact_data(commands))
    result = audit_basic_setting_artifacts(
        *paths, forbidden_values=(PRIVATE_HOSTNAME, PRIVATE_NTP, *PRIVATE_PASSWORDS)
    )
    assert result == {
        "cases": 1,
        "steps": 4,
        "commands": len(commands),
        "artifacts": {
            "json": "basic.json",
            "html": "basic.html",
            "excel": "basic.xlsx",
        },
    }


def test_basic_artifact_audit_reports_sensitive_location_without_value(
    backend_stub, tmp_path
):
    commands = _all_basic_commands(backend_stub)[:1]
    data = _artifact_data(commands)
    secret = "artifact-sensitive-value-Q7x!"
    data["test_cases"][0]["steps"][0]["details"].append(secret)
    paths = _write_artifacts(tmp_path, data)

    with pytest.raises(AssertionError) as exc_info:
        audit_basic_setting_artifacts(*paths, forbidden_values=(secret,))
    message = str(exc_info.value)
    assert secret not in message
    assert "位置" in message
    assert "json" in message and "html" in message and "excel" in message


@pytest.mark.parametrize(
    ("private_value", "expected_label"),
    [
        ("02:11:22:33:44:91", "硬件地址"),
        (r"C:\Users\private-user-Q7x\report", "本机用户路径"),
    ],
)
def test_basic_artifact_audit_rejects_pattern_leaks_without_echoing_value(
    backend_stub, tmp_path, private_value, expected_label
):
    data = _artifact_data(_all_basic_commands(backend_stub)[:1])
    data["test_cases"][0]["steps"][0]["details"].append(private_value)
    paths = _write_artifacts(tmp_path, data)

    with pytest.raises(AssertionError) as exc_info:
        audit_basic_setting_artifacts(*paths)
    message = str(exc_info.value)
    assert private_value not in message
    assert expected_label in message
    assert "位置" in message


def test_basic_artifact_audit_rejects_excel_command_order_mismatch(
    backend_stub, tmp_path
):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    commands = _all_basic_commands(backend_stub)[:2]
    paths = _write_artifacts(tmp_path, _artifact_data(commands))
    workbook = load_workbook(paths[2])
    try:
        sheet = workbook["复验命令"]
        first = sheet.cell(2, 14).value
        second = sheet.cell(3, 14).value
        sheet.cell(2, 14, second)
        sheet.cell(3, 14, first)
        workbook.save(paths[2])
    finally:
        workbook.close()

    with pytest.raises(AssertionError, match="JSON/Excel"):
        audit_basic_setting_artifacts(*paths)
