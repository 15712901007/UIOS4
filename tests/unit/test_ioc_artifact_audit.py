from __future__ import annotations

import json
from copy import deepcopy
from html import escape

import pytest

from utils.ioc_artifact_audit import (
    REQUIRED_PAGES,
    REQUIRED_SCRIPTS,
    REQUIRED_SECTIONS,
    REQUIRED_SUBPAGES,
    audit_ioc_artifacts,
)
from utils.test_results_to_excel import export_results_to_excel


def _command(text: str = "/usr/ikuai/function/ioc_monitor show TYPE=stats time_range=today"):
    return {
        "target": "router",
        "target_label": "路由器",
        "host": "192.0.2.1",
        "shell": "sh",
        "purpose": "读取威胁态势统计",
        "command": text,
        "expected": "返回今日统计字段且命令只读",
        "actual": "已观察到结构化统计响应",
        "effect": "read_only",
        "valid_when": "页面加载完成后、清理前",
        "copy_ready": True,
        "contains_secret": False,
        "interactive": False,
    }


def _details():
    coverage = [
        "页面：" + "、".join(REQUIRED_PAGES),
        "子页面：" + "、".join(REQUIRED_SUBPAGES),
        "底层脚本：" + "、".join(REQUIRED_SCRIPTS),
    ]
    evidence = [f"{section}\n审计夹具中的可复核结果" for section in REQUIRED_SECTIONS]
    return coverage + evidence


def _data(commands=None):
    commands = commands or [_command(), _command(
        "/usr/ikuai/function/ioc_alert show TYPE=stats"
    )]
    step = {
        "name": "步骤1 操作：打开六个页面并读取数据；验证：页面、协议和后端结果一致",
        "description": "威胁情报中心综合审计",
        "status": "passed",
        "duration": "1.00s",
        "details": _details(),
        "verification_commands": commands,
    }
    return {
        "total": 1,
        "passed": 1,
        "failed": 0,
        "skipped": 0,
        "total_steps": 1,
        "test_cases": [{
            "name": "安全中心-威胁情报中心综合测试",
            "original_name": "test_threat_intelligence_comprehensive",
            "status": "passed",
            "duration": "1.00s",
            "step_count": 1,
            "steps": [step],
        }],
    }


def _write_artifacts(tmp_path, data):
    tmp_path.mkdir(parents=True, exist_ok=True)
    json_path = tmp_path / "ioc.json"
    html_path = tmp_path / "ioc.html"
    excel_path = tmp_path / "ioc.xlsx"
    json_path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    commands = [
        item["command"]
        for step in data["test_cases"][0]["steps"]
        for item in step["verification_commands"]
    ]
    html_path.write_text(
        "<html><body>"
        + "".join(
            f'<code id="verification-command-0-0-{index}">{escape(command)}</code>'
            for index, command in enumerate(commands)
        )
        + "</body></html>",
        encoding="utf-8",
    )
    ok, message = export_results_to_excel(str(json_path), str(excel_path))
    assert ok, message
    return json_path, html_path, excel_path


def test_ioc_artifact_audit_accepts_complete_consistent_reports(tmp_path):
    paths = _write_artifacts(tmp_path, _data())

    result = audit_ioc_artifacts(*paths, sensitive_values=["runtime-only-value"])

    assert result == {
        "cases": 1,
        "steps": 1,
        "commands": 2,
        "pages": 6,
        "subpages": 3,
        "scripts": 10,
        "artifacts": {
            "json": "ioc.json",
            "html": "ioc.html",
            "excel": "ioc.xlsx",
        },
    }


@pytest.mark.parametrize("missing_kind", ["section", "page", "subpage", "script"])
def test_ioc_artifact_audit_rejects_incomplete_coverage(tmp_path, missing_kind):
    data = _data()
    details = data["test_cases"][0]["steps"][0]["details"]
    if missing_kind == "section":
        details.remove(next(item for item in details if REQUIRED_SECTIONS[0] in item))
    elif missing_kind == "page":
        # Remove the last page; the first page name also appears in the
        # command purpose (and would therefore still be covered).
        details[0] = "页面：" + "、".join(list(REQUIRED_PAGES)[:-1])
    elif missing_kind == "subpage":
        # Avoid the Syslog alias, which is intentionally accepted in the
        # ioc_syslog script name; remove the Chinese blacklist label instead.
        details[1] = "子页面：" + "、".join(
            [list(REQUIRED_SUBPAGES)[0], list(REQUIRED_SUBPAGES)[2]]
        )
    else:
        details[2] = "底层脚本：" + "、".join(REQUIRED_SCRIPTS[1:])

    paths = _write_artifacts(tmp_path, data)
    with pytest.raises(AssertionError):
        audit_ioc_artifacts(*paths)


@pytest.mark.parametrize("unsafe", [
    "for x in a; do echo $x; done",
    "rm -rf /tmp/ioc-test",
    "password=cleartext",
])
def test_ioc_artifact_audit_rejects_unsafe_commands(tmp_path, unsafe):
    data = _data([_command(unsafe)])
    paths = _write_artifacts(tmp_path, data)

    with pytest.raises(AssertionError):
        audit_ioc_artifacts(*paths)


def test_ioc_artifact_audit_rejects_json_html_command_mismatch(tmp_path):
    data = _data()
    paths = _write_artifacts(tmp_path, data)
    paths[1].write_text(
        paths[1].read_text(encoding="utf-8").replace(
            "TYPE=stats", "TYPE=stats%20tampered", 1
        ),
        encoding="utf-8",
    )

    with pytest.raises(AssertionError):
        audit_ioc_artifacts(*paths)


def test_ioc_artifact_audit_rejects_excel_command_mismatch(tmp_path):
    data = _data()
    paths = _write_artifacts(tmp_path, data)
    from openpyxl import load_workbook

    workbook = load_workbook(paths[2])
    try:
        workbook["复验命令"].cell(2, 14).value = "tampered command"
        workbook.save(paths[2])
    finally:
        workbook.close()

    with pytest.raises(AssertionError):
        audit_ioc_artifacts(*paths)


def test_ioc_artifact_audit_rejects_sensitive_values_and_formula_cells(tmp_path):
    data = _data()
    data["test_cases"][0]["steps"][0]["details"].append(
        "后端原始值：ioc-runtime-secret-1234"
    )
    paths = _write_artifacts(tmp_path, data)
    with pytest.raises(AssertionError):
        audit_ioc_artifacts(*paths, sensitive_values=["ioc-runtime-secret-1234"])

    formula_data = _data()
    formula_data["test_cases"][0]["steps"][0]["details"].append("=SUM(1,1)")
    formula_paths = _write_artifacts(tmp_path / "formula", formula_data)
    with pytest.raises(AssertionError):
        audit_ioc_artifacts(*formula_paths)


def test_ioc_artifact_audit_can_skip_subpage_requirement(tmp_path):
    data = _data()
    data["test_cases"][0]["steps"][0]["details"][1] = "子页面：仅保留页面导航"
    paths = _write_artifacts(tmp_path, data)

    result = audit_ioc_artifacts(*paths, require_subpages=False)

    assert result["subpages"] == 0


def test_ioc_artifact_audit_does_not_mutate_input_data(tmp_path):
    data = _data()
    before = deepcopy(data)
    paths = _write_artifacts(tmp_path, data)
    audit_ioc_artifacts(*paths)
    assert data == before
